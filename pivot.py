import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Load the Excel file
file_path = 'data.xlsx'
data_sheet_name = 'Data'

# Read the data from the "Data" sheet
df = pd.read_excel(file_path, sheet_name=data_sheet_name)

# Preprocess the Labels column to categorize into 'dd-' labels or 'None'
def categorize_labels(label):
    if isinstance(label, str):
        # Split the labels by semicolon and filter out those starting with 'dd-'
        dd_labels = [part.strip() for part in label.split(';') if part.strip().startswith('dd-')]
        return dd_labels[0] if dd_labels else 'None'  # Return the first matching dd- label or 'None'
    return 'None'

# Apply the categorization function to the Labels column
df['Categorized Labels'] = df['Labels'].apply(categorize_labels)

# Create the Pivot Table
pivot_table = pd.pivot_table(
    df,
    index=['Project', 'Custom field (Automation Status/Reason/Solution)'],
    columns='Categorized Labels',
    aggfunc='size',
    fill_value=0
)

# Add a Grand Total column
pivot_table['Grand Total'] = pivot_table.sum(axis=1)

# Reset index to flatten the pivot table
pivot_table_reset = pivot_table.reset_index()

# Write the pivot table to a new sheet in the same Excel file
with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    pivot_table_reset.to_excel(writer, sheet_name='Pivoted Data', index=False)
    
    # Access the workbook and the newly created sheet
    workbook = writer.book
    worksheet = writer.sheets['Pivoted Data']
    
    # Define fill colors
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # Red
    
    # Iterate through rows and apply background colors based on status
    for row in worksheet.iter_rows(min_row=2, max_col=2, max_row=worksheet.max_row):
        status_cell = row[1]  # Assuming status is in the second column
        status = status_cell.value
        
        if status in ["Feasible-Automation In Progress-Python", "Fully Automated-ICFDQ (Data Validation)"]:
            status_cell.fill = green_fill
        elif status in ["Not Feasible-Not Automated", "Partially Automated-ICCDQ (Data Validation)"]:
            status_cell.fill = red_fill
    
    # Add hyperlinks to the numbers in the pivot table
    for row_idx, row_data in enumerate(pivot_table_reset.itertuples(index=False), start=2):  # Start from row 2
        project = row_data[0]  # Project name
        status = row_data[1]   # Automation Status
        
        for col_idx, label in enumerate(pivot_table_reset.columns[2:], start=3):  # Start from column 3 (labels)
            count = row_data[col_idx - 1]  # Get the count value
            
            if count > 0:  # Only add hyperlink if count > 0
                # Filter the original DataFrame to get matching rows
                filtered_df = df[(df['Project'] == project) & 
                                 (df['Custom field (Automation Status/Reason/Solution)'] == status) &
                                 (df['Categorized Labels'] == label)]
                
                if not filtered_df.empty:
                    # Create a hyperlink to the first matching row in the "Data" sheet
                    first_row = filtered_df.index[0] + 2  # Add 2 because Excel rows are 1-indexed and header is row 1
                    hyperlink = f"{file_path}#Data!A{first_row}"
                    
                    # Set the hyperlink and make the cell clickable
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.hyperlink = hyperlink
                    cell.value = count  # Ensure the cell value is the count
                    cell.style = "Hyperlink"  # Apply the hyperlink style

print(f"Pivot table has been added to the sheet 'Pivoted Data' in {file_path} with colored statuses and hyperlinks.")
