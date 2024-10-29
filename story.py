import requests
from requests_kerberos import HTTPKerberosAuth, REQUIRED
import concurrent.futures
import time
import conf  # Import the configuration file
import logging
import csv
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Jira server URL
jira_url = 'https://your-company-jira.com'

# Kerberos authentication
kerberos_auth = HTTPKerberosAuth(mutual_authentication=REQUIRED)

# Function to fetch all issues for a given JQL query
def fetch_all_issues(jql_query):
    start_at = 0
    max_results = 1000
    all_issues = []
    retry_attempts = 3
    retry_delay = 5

    while True:
        params = {
            'jql': jql_query,
            'startAt': start_at,
            'maxResults': max_results,
            'fields': 'issuetype,priority,versions,components,labels,status,resolution,fixVersions,customfield_12345,customfield_67890,customfield_112233,storyPoints'  # Add all relevant fields here
        }

        for attempt in range(retry_attempts):
            try:
                response = requests.get(f'{jira_url}/rest/api/2/search', params=params, auth=kerberos_auth)
                if response.status_code == 200:
                    issues = response.json()['issues']
                    all_issues.extend(issues)
                    if len(issues) < max_results:
                        return all_issues
                    start_at += max_results
                    break
                else:
                    logging.error(f"Failed to retrieve issues for JQL '{jql_query}': {response.status_code} - {response.text}")
                    if response.status_code == 401:  # Unauthorized
                        logging.warning("Authentication issue, retrying...")
                        time.sleep(retry_delay)
            except requests.exceptions.RequestException as e:
                logging.error(f"Request failed: {e}")
                time.sleep(retry_delay)

        if attempt == retry_attempts - 1:
            logging.error(f"Failed to retrieve issues for JQL '{jql_query}' after {retry_attempts} attempts.")
            break

    return all_issues

# Function to process and categorize issues for email content
def process_issues(jql_query, all_email_content):
    issues = fetch_all_issues(jql_query)
    component_name = jql_query.split('component = "')[1].split('"')[0]

    email_content = f"Results for component {component_name}:<br>"

    for issue in issues:
        issue_key = issue['key']
        issue_summary = issue['fields']['summary']
        issue_type = issue['fields']['issuetype']['name']
        priority = issue['fields'].get('priority', {}).get('name', 'N/A')
        affect_version = ', '.join([ver['name'] for ver in issue['fields'].get('versions', [])])
        component = ', '.join([comp['name'] for comp in issue['fields'].get('components', [])])
        labels = ', '.join(issue['fields'].get('labels', []))
        status = issue['fields']['status']['name']
        resolution = issue['fields'].get('resolution', {}).get('name', 'Unresolved')
        fix_version = ', '.join([ver['name'] for ver in issue['fields'].get('fixVersions', [])])
        qa_required = issue['fields'].get('customfield_12345', 'N/A')  # Replace with actual field ID for QA Required
        sprint = issue['fields'].get('customfield_67890', 'N/A')  # Replace with actual field ID for Sprint
        story_points = issue['fields'].get('customfield_112233', 'N/A')  # Replace with actual field ID for Story Points
        req_status = issue['fields'].get('customfield_445566', 'N/A')  # Replace with actual field ID for Requirement Status

        email_content += f"<br><strong>{issue_key} - {issue_summary}</strong><br>"
        email_content += f"Type: {issue_type}, Priority: {priority}, Affect Version: {affect_version}, Component: {component}, Labels: {labels}, Status: {status}, Resolution: {resolution}, Fix Version: {fix_version}, QA Required?: {qa_required}, Sprint: {sprint}, Story Points: {story_points}, Requirement Status: {req_status}<br>"

    all_email_content.append(email_content)

# Function to send an email with the provided subject and body
def send_email(subject, body):
    sender_email = "your_email@example.com"
    recipient_email = "recipient@example.com"

    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = recipient_email
    msg['Subject'] = subject

    msg.attach(MIMEText(body, 'html'))

    try:
        server = smtplib.SMTP('smtp.example.com', 587)
        server.starttls()
        server.login(sender_email, "your_password")
        server.sendmail(sender_email, recipient_email, msg.as_string())
        server.quit()
        logging.info(f"Email sent to {recipient_email}")
    except Exception as e:
        logging.error(f"Failed to send email: {e}")

# Function to process all JQL queries concurrently and send a combined email
def process_all_jql_queries():
    all_email_content = []  # Accumulate all email content across JQL queries

    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = {executor.submit(process_issues, jql_query, all_email_content): jql_query for jql_query in jql_queries}
        for future in concurrent.futures.as_completed(futures):
            jql_query = futures[future]
            try:
                future.result()
            except Exception as e:
                logging.error(f"Error processing JQL '{jql_query}': {e}")
                retry_attempts = 3
                for attempt in range(retry_attempts):
                    logging.info(f"Retrying JQL '{jql_query}' (Attempt {attempt + 1}/{retry_attempts})")
                    try:
                        process_issues(jql_query, all_email_content)
                        break
                    except Exception as e:
                        logging.error(f"Retry {attempt + 1} failed for JQL '{jql_query}': {e}")
                        time.sleep(5)  # Wait before retrying

    combined_email_content = "\n".join(all_email_content)
    send_email("Jira Report for All Components", combined_email_content)

# Combine components from conf.py into jql_queries
jql_queries = [
    f'created >= "2024-07-31" AND created < "2024-08-01" AND component = "{component}"'
    for component in (conf.components + conf.components1)
]

# Start processing all JQL queries
process_all_jql_queries()



# Below code is putting the data into table instead of plain printing in above code



import requests
from requests_kerberos import HTTPKerberosAuth, REQUIRED
import concurrent.futures
import time
import conf  # Import the configuration file
import logging
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Jira server URL
jira_url = 'https://your-company-jira.com'

# Kerberos authentication
kerberos_auth = HTTPKerberosAuth(mutual_authentication=REQUIRED)

def fetch_all_issues(jql_query):
    start_at = 0
    max_results = 1000
    all_issues = []
    retry_attempts = 3
    retry_delay = 5

    while True:
        params = {
            'jql': jql_query,
            'startAt': start_at,
            'maxResults': max_results
        }

        for attempt in range(retry_attempts):
            try:
                response = requests.get(f'{jira_url}/rest/api/2/search', params=params, auth=kerberos_auth)

                if response.status_code == 200:
                    issues = response.json()['issues']
                    all_issues.extend(issues)
                    if len(issues) < max_results:
                        return all_issues
                    start_at += max_results
                    break
                else:
                    logging.error(f"Failed to retrieve issues for JQL '{jql_query}': {response.status_code} - {response.text}")
                    if response.status_code == 401:  # Unauthorized
                        logging.warning("Authentication issue, retrying...")
                        time.sleep(retry_delay)
            except requests.exceptions.RequestException as e:
                logging.error(f"Request failed: {e}")
                time.sleep(retry_delay)

        if attempt == retry_attempts - 1:
            logging.error(f"Failed to retrieve issues for JQL '{jql_query}' after {retry_attempts} attempts.")
            break

    return all_issues

def generate_html_table(issues, fields):
    table_header = "<tr><th>Serial No</th><th>Story</th><th>Summary</th>"
    for field in fields:
        table_header += f"<th>{field.replace('_', ' ').title()}</th>"
    table_header += "</tr>"

    table_rows = ""
    for i, issue in enumerate(issues, start=1):
        table_row = f"<tr><td>{i}</td><td>{issue['key']}</td><td>{issue['fields']['summary']}</td>"
        for field in fields:
            value = issue['fields'].get(field, "")
            if isinstance(value, dict) and 'name' in value:
                value = value['name']
            elif isinstance(value, list):
                value = ', '.join(str(v['name'] if isinstance(v, dict) and 'name' in v else v) for v in value)
            table_row += f"<td>{value}</td>"
        table_row += "</tr>"
        table_rows += table_row

    html_table = f"<table border='1' cellpadding='5' cellspacing='0'>{table_header}{table_rows}</table>"
    return html_table

def process_issues(jql_query, all_email_content, fields):
    issues = fetch_all_issues(jql_query)
    if not issues:
        return

    component_name = jql_query.split('component = ')[1].split()[0].replace('"', '')
    email_content = f"<h2>Results for component: {component_name}</h2><br>"

    email_content += generate_html_table(issues, fields)

    all_email_content.append(email_content)

def send_email(subject, body):
    sender_email = "your_email@example.com"
    recipient_email = "recipient@example.com"

    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = recipient_email
    msg['Subject'] = subject

    msg.attach(MIMEText(body, 'html'))

    try:
        server = smtplib.SMTP('smtp.example.com', 587)
        server.starttls()
        server.login(sender_email, "your_password")
        server.sendmail(sender_email, recipient_email, msg.as_string())
        server.quit()
        logging.info(f"Email sent to {recipient_email}")
    except Exception as e:
        logging.error(f"Failed to send email: {e}")

def process_all_jql_queries():
    all_email_content = []

    fields = [
        'issuetype',
        'priority',
        'versions',
        'components',
        'labels',
        'status',
        'resolution',
        'fixVersions',
        'customfield_10021',  # QA Required? (custom field)
        'customfield_10020',  # Sprint (custom field)
        'customfield_10002',  # Story Points (custom field)
        'customfield_10010'   # Requirement Status (custom field)
    ]

    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = {executor.submit(process_issues, jql_query, all_email_content, fields): jql_query for jql_query in jql_queries}
        for future in concurrent.futures.as_completed(futures):
            jql_query = futures[future]
            try:
                future.result()
            except Exception as e:
                logging.error(f"Error processing JQL '{jql_query}': {e}")
                retry_attempts = 3
                for attempt in range(retry_attempts):
                    logging.info(f"Retrying JQL '{jql_query}' (Attempt {attempt + 1}/{retry_attempts})")
                    try:
                        process_issues(jql_query, all_email_content, fields)
                        break
                    except Exception as e:
                        logging.error(f"Retry {attempt + 1} failed for JQL '{jql_query}': {e}")
                        time.sleep(5)  # Wait before retrying

    combined_email_content = "<br><br>".join(all_email_content)
    send_email("Jira Report for All Components", combined_email_content)

# Combine components from conf.py into jql_queries
jql_queries = [
    f'created >= "2024-07-31" AND created < "2024-08-01" AND component = "{component}"' 
    for component in (conf.components + conf.components1)
]

# Start processing all JQL queries
process_all_jql_queries()


############## to handle specific fields
    def generate_html_table(issues, fields):
    table_header = "<tr><th>Serial No</th><th>Story</th><th>Summary</th>"
    for field in fields:
        table_header += f"<th>{field.replace('_', ' ').title()}</th>"
    table_header += "</tr>"

    table_rows = ""
    for i, issue in enumerate(issues, start=1):
        table_row = f"<tr><td>{i}</td><td>{issue['key']}</td><td>{issue['fields']['summary']}</td>"
        for field in fields:
            value = issue['fields'].get(field, "")

            # Handle customfield_10005 (Sprint Name)
            if field == 'customfield_10005' and isinstance(value, list) and value:
                value = value[0].split("name=")[-1].split(",")[0]  # Extract name from string

            # Handle customfield_26424 (Status)
            elif field == 'customfield_26424' and isinstance(value, list) and value:
                value = value[0].get('status', '')  # Extract status from dictionary

            elif isinstance(value, dict) and 'name' in value:
                value = value['name']
            elif isinstance(value, list):
                value = ', '.join(str(v['name'] if isinstance(v, dict) and 'name' in v else v) for v in value)

            table_row += f"<td>{value}</td>"
        table_row += "</tr>"
        table_rows += table_row

    html_table = f"<table border='1' cellpadding='5' cellspacing='0'>{table_header}{table_rows}</table>"
    return html_table


# rows with color

import html

# Mapping of custom field IDs to user-friendly names
custom_field_mapping = {
    'customfield_10021': 'QA Required?',
    'customfield_10020': 'Sprint',
    'customfield_10002': 'Story Points',
    'customfield_10010': 'Requirement Status',
    # Add more mappings if needed
}



def generate_html_table(issues, fields):
    # Table header
    table_header = "<tr><th>Serial No</th><th>Story</th><th>Summary</th>"
    for field in fields:
        # Replace custom field IDs with user-friendly names if available
        field_name = custom_field_mapping.get(field, field.replace('_', ' ').title())
        table_header += f"<th>{html.escape(field_name)}</th>"
    table_header += "</tr>"

    table_rows = ""
    for i, issue in enumerate(issues, start=1):
        # Alternate row color: light gray for odd rows, white for even rows
        row_color = "#f2f2f2" if i % 2 != 0 else "#ffffff"
        table_row = f"<tr style='background-color:{row_color};'><td>{i}</td><td>{html.escape(issue['key'])}</td><td>{html.escape(issue['fields']['summary'])}</td>"

        for field in fields:
            value = issue['fields'].get(field, "")

            # Handle customfield_10005 (Sprint Name)
            if field == 'customfield_10005' and isinstance(value, list) and value:
                value = value[0].split("name=")[-1].split(",")[0]  # Extract name from string

            # Handle customfield_26424 (Status)
            elif field == 'customfield_26424' and isinstance(value, list) and value:
                value = value[0].get('status', '')  # Extract status from dictionary

            elif isinstance(value, dict) and 'name' in value:
                value = value['name']
            elif isinstance(value, list):
                value = ', '.join(str(v['name'] if isinstance(v, dict) and 'name' in v else v) for v in value)

            # Escape the cell content to prevent HTML parsing issues
            table_row += f"<td>{html.escape(str(value))}</td>"

        table_row += "</tr>"
        table_rows += table_row

    # Add CSS for table layout consistency
    html_table = f"""
    <table border='1' cellpadding='5' cellspacing='0' style='border-collapse: collapse; width: 100%;'>
        {table_header}
        {table_rows}
    </table>
    """
    return html_table


### predictive JIRA complican analysis

def generate_html_table(issues, fields):
    # Table header
    table_header = "<tr><th>Serial No</th><th>Story</th><th>Summary</th>"
    for field in fields:
        field_name = custom_field_mapping.get(field, field.replace('_', ' ').title())
        table_header += f"<th>{html.escape(field_name)}</th>"
    table_header += "</tr>"

    table_rows = ""
    for i, issue in enumerate(issues, start=1):
        row_color = "#f2f2f2" if i % 2 != 0 else "#ffffff"
        table_row = f"<tr style='background-color:{row_color};'><td>{i}</td><td>{html.escape(issue['key'])}</td><td>{html.escape(issue['fields']['summary'])}</td>"


        for field in fields:
            value = issue['fields'].get(field, "")

            if field == 'customfield_10005' and isinstance(value, list) and value:
                value = value[0].split("name=")[-1].split(",")[0]

            elif field == 'customfield_26424' and isinstance(value, list) and value:
                value = value[0].get('status', '')

            elif isinstance(value, dict) and 'name' in value:
                value = value['name']
            elif isinstance(value, list):
                value = ', '.join(str(v['name'] if isinstance(v, dict) and 'name' in v else v) for v in value)

            cell_style = ""
            if qa_assignee and field == 'customfield_17201' and qa_required != 'Yes':
                cell_style = "background-color: #ffcccc;"  # Light red color for QA Required? cell
            elif qa_assignee and field == 'customfield_26424' and requirement_status != 'Ok':
                cell_style = "background-color: #ffcccc;"  # Light red color for Requirement Status cell

            # Escape the cell content to prevent HTML parsing issues
            table_row += f"<td>{html.escape(str(value))}</td>"

        table_row += "</tr>"
        table_rows += table_row

    html_table = f"""
    <table border='1' cellpadding='5' cellspacing='0' style='border-collapse: collapse; width: 100%;'>
        {table_header}
        {table_rows}
    </table>
    """
    return html_table

##### highligting colot


import requests
from requests_kerberos import HTTPKerberosAuth, REQUIRED
import concurrent.futures
import time
import conf  # Import the configuration file
import logging
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from bs4 import BeautifulSoup  # Import BeautifulSoup for modifying HTML content

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Jira server URL
jira_url = 'https://your-company-jira.com'

# Kerberos authentication
kerberos_auth = HTTPKerberosAuth(mutual_authentication=REQUIRED)

# Mapping of custom field IDs to user-friendly names
custom_field_mapping = {
    'customfield_17201': 'QA Required?',
    'customfield_10005': 'Sprint',
    'customfield_10002': 'Story Points',
    'customfield_26424': 'Requirement Status',
    'customfield_26027': 'QA Assignee',
    # Add more mappings if needed
}

def fetch_all_issues(jql_query):
    start_at = 0
    max_results = 1000
    all_issues = []
    retry_attempts = 3
    retry_delay = 5

    while True:
        params = {
            'jql': jql_query,
            'startAt': start_at,
            'maxResults': max_results
        }

        for attempt in range(retry_attempts):
            try:
                response = requests.get(f'{jira_url}/rest/api/2/search', params=params, auth=kerberos_auth)

                if response.status_code == 200:
                    issues = response.json()['issues']
                    all_issues.extend(issues)
                    if len(issues) < max_results:
                        return all_issues
                    start_at += max_results
                    break
                else:
                    logging.error(f"Failed to retrieve issues for JQL '{jql_query}': {response.status_code} - {response.text}")
                    if response.status_code == 401:  # Unauthorized
                        logging.warning("Authentication issue, retrying...")
                        time.sleep(retry_delay)
            except requests.exceptions.RequestException as e:
                logging.error(f"Request failed: {e}")
                time.sleep(retry_delay)

        if attempt == retry_attempts - 1:
            logging.error(f"Failed to retrieve issues for JQL '{jql_query}' after {retry_attempts} attempts.")
            break

    return all_issues

def generate_html_table(issues, fields):
    # Table header
    table_header = "<tr><th>Serial No</th><th>Story</th><th>Summary</th>"
    for field in fields:
        # Replace custom field IDs with user-friendly names if available
        field_name = custom_field_mapping.get(field, field.replace('_', ' ').title())
        table_header += f"<th>{html.escape(field_name)}</th>"
    table_header += "</tr>"

    table_rows = []
    for i, issue in enumerate(issues, start=1):
        # Alternate row color: light gray for odd rows, white for even rows
        row_color = "#f2f2f2" if i % 2 != 0 else "#ffffff"
        table_row = f"<tr style='background-color:{row_color};'><td>{i}</td><td>{html.escape(issue['key'])}</td><td>{html.escape(issue['fields']['summary'])}</td>"

        row_cells = []
        for field in fields:
            value = issue['fields'].get(field, "")

            # Handle customfield_10005 (Sprint Name)
            if field == 'customfield_10005' and isinstance(value, list) and value:
                value = value[0].split("name=")[-1].split(",")[0]  # Extract name from string

            # Handle customfield_26424 (Status)
            elif field == 'customfield_26424' and isinstance(value, list) and value:
                value = value[0].get('status', '')  # Extract status from dictionary

            elif isinstance(value, dict) and 'name' in value:
                value = value['name']
            elif isinstance(value, list):
                value = ', '.join(str(v['name'] if isinstance(v, dict) and 'name' in v else v) for v in value)

            # Escape the cell content to prevent HTML parsing issues
            row_cells.append(f"<td>{html.escape(str(value))}</td>")

        table_row += "".join(row_cells) + "</tr>"
        table_rows.append(table_row)

    # Add CSS for table layout consistency
    html_table = f"""
    <table border='1' cellpadding='5' cellspacing='0' style='border-collapse: collapse; width: 100%;'>
        {table_header}
        {''.join(table_rows)}
    </table>
    """
    return html_table

def apply_rules_to_table(html_table, issues, fields):
    soup = BeautifulSoup(html_table, 'html.parser')

    for i, issue in enumerate(issues):
        qa_assignee = issue['fields'].get('customfield_26027', None)  # QA Assignee

        if qa_assignee:  # If QA Assignee is not None
            table_row = soup.find_all('tr')[i + 1]  # +1 because the first row is the header

            # QA Required? column check
            qa_required = issue['fields'].get('customfield_17201', "")
            if qa_required != "Yes":
                qa_required_cell = table_row.find_all('td')[fields.index('customfield_17201') + 3]  # Adjust for index
                qa_required_cell['style'] = 'background-color: #ffcccc;'  # Light red color

            # Requirement Status column check
            requirement_status = ""
            customfield_26424_value = issue['fields'].get('customfield_26424', [])
            if isinstance(customfield_26424_value, list) and customfield_26424_value:
                requirement_status = customfield_26424_value[0].get('status', "")
            if requirement_status != "Ok":
                requirement_status_cell = table_row.find_all('td')[fields.index('customfield_26424') + 3]
                requirement_status_cell['style'] = 'background-color: #ffcccc;'  # Light red color

    return str(soup)

def process_issues(jql_query, all_email_content, fields):
    issues = fetch_all_issues(jql_query)
    if not issues:
        return

    component_name = jql_query.split('component = ')[1].split()[0].replace('"', '')
    email_content = f"<h2>Results for component: {component_name}</h2><br>"

    table_html = generate_html_table(issues, fields)
    table_html_with_rules = apply_rules_to_table(table_html, issues, fields)
    
    email_content += table_html_with_rules

    all_email_content.append(email_content)

def send_email(subject, body):
    sender_email = "your_email@example.com"
    recipient_email = "recipient@example.com"

    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = recipient_email
    msg['Subject'] = subject

    msg.attach(MIMEText(body, 'html'))

    try:
        server = smtplib.SMTP('smtp.example.com', 587)
        server.starttls()
        server.login(sender_email, "your_password")
        server.sendmail(sender_email, recipient_email, msg.as_string())
        server.quit()
        logging.info(f"Email sent to {recipient_email}")
    except Exception as e:
        logging.error(f"Failed to send email: {e}")

def process_all_jql_queries():
    all_email_content = []

    fields = [
        'issuetype',
        'priority',
        'versions',
        'components',
        'labels',
        'status',
        'resolution',
        'fixVersions',
        'customfield_17201',  # QA Required?
        'customfield_10005',   # Sprint
        'customfield_10002',   # Story Points
        'customfield_26424',   # Requirement Status
        'customfield_26027',   # QA Assignee
    ]

    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = {executor.submit(process_issues, jql_query, all_email_content, fields): jql_query for jql_query in jql_queries}
        for future in concurrent.futures.as_completed(futures):
            jql_query = futures[future]
            try:
                future.result()
            except Exception as e:
                logging.error(f"Error processing JQL '{jql_query}': {e}")
                retry_attempts = 3
                for attempt in range(retry_attempts):
                    logging.info(f"Retrying JQL '{jql_query}' (Attempt {attempt + 1}/{retry_attempts})")
                    try:
                        process_issues(jql_query, all_email_content, fields)
                        break
                    except Exception as e:
                        logging.error(f"Retry {attempt + 1} failed for JQL '{jql_query}': {e}")
                        time.sleep(5)  # Wait before retrying

    combined_email_content = "<br><br>".join(all_email_content)


########################33 subtasks fetch

import requests
from requests_kerberos import HTTPKerberosAuth, REQUIRED
import concurrent.futures
import time
import html
import logging
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Jira server URL
jira_url = 'https://your-company-jira.com'

# Kerberos authentication
kerberos_auth = HTTPKerberosAuth(mutual_authentication=REQUIRED)

# Function to fetch a single subtask
def fetch_subtask(subtask_key):
    url = f'{jira_url}/rest/api/2/issue/{subtask_key}'
    try:
        response = requests.get(url, auth=kerberos_auth)
        if response.status_code == 200:
            return response.json()
        else:
            logging.error(f"Failed to retrieve subtask '{subtask_key}': {response.status_code} - {response.text}")
    except requests.exceptions.RequestException as e:
        logging.error(f"Request failed: {e}")
    return None

# Function to fetch subtasks for a list of keys
def fetch_subtasks(subtask_keys):
    subtasks = []
    for key in subtask_keys:
        subtask = fetch_subtask(key)
        if subtask:
            subtasks.append(subtask)
    return subtasks

# Mapping of custom field IDs to user-friendly names
custom_field_mapping = {
    'customfield_17201': 'QA Required?',
    'customfield_10005': 'Sprint',
    'customfield_10002': 'Story Points',
    'customfield_26424': 'Requirement Status',
    'customfield_26027': 'QA Assignee',
    # Add more mappings if needed
}

# Modified function to generate the HTML table with subtasks
def generate_html_table(issues, fields):
    # Table header
    table_header = "<tr><th>Serial No</th><th>Story</th><th>Summary</th>"
    for field in fields:
        field_name = custom_field_mapping.get(field, field.replace('_', ' ').title())
        table_header += f"<th>{html.escape(field_name)}</th>"
    table_header += "<th>Subtasks</th></tr>"

    table_rows = ""
    for i, issue in enumerate(issues, start=1):
        row_color = "#f2f2f2" if i % 2 != 0 else "#ffffff"
        table_row = f"<tr style='background-color:{row_color};'><td>{i}</td><td>{html.escape(issue['key'])}</td><td>{html.escape(issue['fields']['summary'])}</td>"

        for field in fields:
            value = issue['fields'].get(field, "")
            if field == 'customfield_10005' and isinstance(value, list) and value:
                value = value[0].split("name=")[-1].split(",")[0]
            elif field == 'customfield_26424' and isinstance(value, list) and value:
                value = value[0].get('status', '')
            elif isinstance(value, dict) and 'name' in value:
                value = value['name']
            elif isinstance(value, list):
                value = ', '.join(str(v['name'] if isinstance(v, dict) and 'name' in v else v) for v in value)
            table_row += f"<td>{html.escape(str(value))}</td>"

        # Fetch subtasks and add them to the table row
        subtask_keys = [subtask['key'] for subtask in issue['fields'].get('subtasks', [])]
        if subtask_keys:
            subtasks = fetch_subtasks(subtask_keys)
            subtask_summaries = ', '.join([html.escape(subtask['fields']['summary']) for subtask in subtasks])
            table_row += f"<td>{subtask_summaries}</td>"
        else:
            table_row += "<td>No Subtasks</td>"

        table_row += "</tr>"
        table_rows += table_row

    html_table = f"""
    <table border='1' cellpadding='5' cellspacing='0' style='border-collapse: collapse; width: 100%;'>
        {table_header}
        {table_rows}
    </table>
    """
    return html_table

def process_issues(jql_query, all_email_content, fields):
    issues = fetch_all_issues(jql_query)
    if not issues:
        return

    component_name = jql_query.split('component = ')[1].split()[0].replace('"', '')
    email_content = f"<h2>Results for component: {component_name}</h2><br>"

    email_content += generate_html_table(issues, fields)

    all_email_content.append(email_content)

def send_email(subject, body):
    sender_email = "your_email@example.com"
    recipient_email = "recipient@example.com"

    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = recipient_email
    msg['Subject'] = subject

    msg.attach(MIMEText(body, 'html'))

    try:
        server = smtplib.SMTP('smtp.example.com', 587)
        server.starttls()
        server.login(sender_email, "your_password")
        server.sendmail(sender_email, recipient_email, msg.as_string())
        server.quit()
        logging.info(f"Email sent to {recipient_email}")
    except Exception as e:
        logging.error(f"Failed to send email: {e}")

def process_all_jql_queries():
    all_email_content = []

    fields = [
        'issuetype',
        'priority',
        'versions',
        'components',
        'labels',
        'status',
        'resolution',
        'fixVersions',
        'customfield_17201',
        'customfield_10005',
        'customfield_10002',
        'customfield_26424',
        'customfield_26027',
    ]

    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = {executor.submit(process_issues, jql_query, all_email_content, fields): jql_query for jql_query in jql_queries}
        for future in concurrent.futures.as_completed(futures):
            jql_query = futures[future]
            try:
                future.result()
            except Exception as e:
                logging.error(f"Error processing JQL '{jql_query}': {e}")
                retry_attempts = 3
                for attempt in range(retry_attempts):
                    logging.info(f"Retrying JQL '{jql_query}' (Attempt {attempt + 1}/{retry_attempts})")
                    try:
                        process_issues(jql_query, all_email_content, fields)
                        break
                    except Exception as e:
                        logging.error(f"Retry {attempt + 1} failed for JQL '{jql_query}': {e}")
                        time.sleep(5)

    combined_email_content = "<br><br>".join(all_email_content)
    send_email("Jira Report for All Components", combined_email_content)

# Combine components from conf.py into jql_queries
jql_queries = [
    f'created >= "2024-07-31" AND created < "2024-08-01" AND component = "{component}"' 
    for component in (conf.components + conf.components1)
]

# Start processing all JQL queries
process_all_jql_queries()



#### coloring

def generate_html_table(issues, fields):
    # Table header
    table_header = "<tr><th>Serial No</th><th>Story</th><th>Summary</th>"
    for field in fields:
        # Replace custom field IDs with user-friendly names if available
        field_name = custom_field_mapping.get(field, field.replace('_', ' ').title())
        table_header += f"<th>{html.escape(field_name)}</th>"
    table_header += "</tr>"

    table_rows = ""
    for i, issue in enumerate(issues, start=1):
        # Alternate row color: light gray for odd rows, white for even rows
        row_color = "#f2f2f2" if i % 2 != 0 else "#ffffff"
        table_row = f"<tr style='background-color:{row_color};'><td>{i}</td><td>{html.escape(issue['key'])}</td><td>{html.escape(issue['fields']['summary'])}</td>"

        qa_required = None
        requirement_status = None

        for field in fields:
            if field == 'subtasks':
                subtasks = issue['fields'].get('subtasks', [])
                subtask_keys = [subtask['key'] for subtask in subtasks]
                value = ', '.join(subtask_keys)
            else:
                value = issue['fields'].get(field, "")

            # Handle customfield_10005 (Sprint Name)
            if field == 'customfield_10005' and isinstance(value, list) and value:
                value = value[0].split("name=")[-1].split(",")[0]  # Extract name from string

            # Handle customfield_26424 (Requirement Status)
            elif field == 'customfield_26424' and isinstance(value, list) and value:
                value = value[0].get('status', '')  # Extract status from dictionary
                requirement_status = value  # Capture Requirement Status for later use

            elif isinstance(value, dict) and 'name' in value:
                value = value['name']
            elif isinstance(value, list):
                value = ', '.join(str(v['name'] if isinstance(v, dict) and 'name' in v else v) for v in value)

            # Capture QA Required? for logic check
            if field == 'customfield_17201':
                qa_required = value

            # Escape the cell content to prevent HTML parsing issues
            cell_content = html.escape(str(value))

            # Apply the highlighting rules
            if field == 'customfield_26424' and requirement_status is not None and qa_required is not None:
                if (qa_required == "Yes" and requirement_status != "OK") or (qa_required == "No" and requirement_status == "OK"):
                    table_row += f"<td style='color: red;'>{cell_content}</td>"
                else:
                    table_row += f"<td>{cell_content}</td>"
            else:
                table_row += f"<td>{cell_content}</td>"

        table_row += "</tr>"
        table_rows += table_row

    # Add CSS for table layout consistency
    html_table = f"""
    <table border='1' cellpadding='5' cellspacing='0' style='border-collapse: collapse; width: 100%;'>
        {table_header}
        {table_rows}
    </table>
    """
    return html_table


..........................
def generate_html_table(issues, fields):
    # Table header
    table_header = "<tr><th>Serial No</th><th>Story</th><th>Summary</th>"
    for field in fields:
        # Replace custom field IDs with user-friendly names if available
        field_name = custom_field_mapping.get(field, field.replace('_', ' ').title())
        table_header += f"<th>{html.escape(field_name)}</th>"
    table_header += "</tr>"

    table_rows = ""
    for i, issue in enumerate(issues, start=1):
        # Alternate row color: light gray for odd rows, white for even rows
        row_color = "#f2f2f2" if i % 2 != 0 else "#ffffff"
        table_row = f"<tr style='background-color:{row_color};'><td>{i}</td><td>{html.escape(issue['key'])}</td><td>{html.escape(issue['fields']['summary'])}</td>"

        qa_required = None
        requirement_status = None

        for field in fields:
            if field == 'subtasks':
                subtasks = issue['fields'].get('subtasks', [])
                subtask_keys = [subtask['key'] for subtask in subtasks]
                value = ', '.join(subtask_keys)
            else:
                value = issue['fields'].get(field, "")

            # Handle customfield_10005 (Sprint Name)
            if field == 'customfield_10005' and isinstance(value, list) and value:
                value = value[0].split("name=")[-1].split(",")[0]  # Extract name from string

            # Handle customfield_26424 (Requirement Status)
            elif field == 'customfield_26424' and isinstance(value, list) and value:
                value = value[0].get('status', '')  # Extract status from dictionary
                requirement_status = value  # Capture Requirement Status for later use

            elif isinstance(value, dict) and 'name' in value:
                value = value['name']
            elif isinstance(value, list):
                value = ', '.join(str(v['name'] if isinstance(v, dict) and 'name' in v else v) for v in value)

            # Capture QA Required? for logic check and clean it up
            if field == 'customfield_17201':
                qa_required = str(value).replace("!", "").strip()  # Remove '!' and trim whitespace

            # Escape the cell content to prevent HTML parsing issues
            cell_content = html.escape(str(value))

            # Apply the highlighting rules
            if field == 'customfield_26424' and requirement_status is not None:
                if qa_required is None:
                    # Case where QA Required? is None
                    if requirement_status == "OK":
                        table_row += f"<td style='color: red;'>{cell_content}</td>"
                    else:
                        table_row += f"<td>{cell_content}</td>"
                else:
                    # Apply the rules based on the cleaned up QA Required? value
                    if (qa_required == "Yes" and requirement_status != "OK") or (qa_required == "No" and requirement_status == "OK"):
                        table_row += f"<td style='color: red;'>{cell_content}</td>"
                    else:
                        table_row += f"<td>{cell_content}</td>"
            else:
                table_row += f"<td>{cell_content}</td>"

        table_row += "</tr>"
        table_rows += table_row

    # Add CSS for table layout consistency
    html_table = f"""
    <table border='1' cellpadding='5' cellspacing='0' style='border-collapse: collapse; width: 100%;'>
        {table_header}
        {table_rows}
    </table>
    """
    return html_table

########################################################################################## final

import html

def generate_html_table(issues, fields):
    # Table header
    table_header = "<tr><th>Serial No</th><th>Story</th><th>Summary</th>"
    for field in fields:
        # Replace custom field IDs with user-friendly names if available
        field_name = custom_field_mapping.get(field, field.replace('_', ' ').title())
        table_header += f"<th>{html.escape(field_name)}</th>"
    table_header += "</tr>"

    # Define column widths for fixed table layout
    colgroup = """
    <colgroup>
        <col style="width: 5%;">
        <col style="width: 15%;">
        <col style="width: 20%;">
        {col_widths}
    </colgroup>
    """.format(col_widths="".join(['<col style="width: 10%;">' for _ in fields]))

    table_rows = ""
    for i, issue in enumerate(issues, start=1):
        # Alternate row color: light gray for odd rows, white for even rows
        row_color = "#f2f2f2" if i % 2 != 0 else "#ffffff"
        table_row = f"<tr style='background-color:{row_color};'><td>{i}</td><td>{html.escape(issue['key'])}</td><td>{html.escape(issue['fields']['summary'])}</td>"

        for field in fields:
            if field == 'subtasks':
                subtasks = issue['fields'].get('subtasks', [])
                subtask_keys = [subtask['key'] for subtask in subtasks]
                value = ', '.join(subtask_keys)
            else:
                value = issue['fields'].get(field, "")

            # Handle customfield_10005 (Sprint Name)
            if field == 'customfield_10005' and isinstance(value, list) and value:
                value = value[0].split("name=")[-1].split(",")[0]  # Extract name from string

            # Handle customfield_26424 (Status)
            elif field == 'customfield_26424' and isinstance(value, list) and value:
                value = value[0].get('status', '')  # Extract status from dictionary

            elif isinstance(value, dict) and 'name' in value:
                value = value['name']
            elif isinstance(value, list):
                value = ', '.join(str(v['name'] if isinstance(v, dict) and 'name' in v else v) for v in value)

            # Replace None with an empty string to prevent merging issues
            if value is None:
                value = ""

            # Replace any "!" character in the value
            if isinstance(value, str):
                value = value.replace("!", "")

            # Escape the cell content to prevent HTML parsing issues
            table_row += f"<td>{html.escape(str(value))}</td>"

        table_row += "</tr>"
        table_rows += table_row

    # Add CSS for table layout consistency
    html_table = f"""
    <table border='1' cellpadding='5' cellspacing='0' style='border-collapse: collapse; width: 100%; table-layout: fixed;'>
        {colgroup}
        {table_header}
        {table_rows}
    </table>
    """
    return html_table


##########################33 try this

def generate_html_table(issues, fields):
    # Table header
    table_header = "<tr><th>Serial No</th><th>Story</th><th>Summary</th>"
    for field in fields:
        # Replace custom field IDs with user-friendly names if available
        field_name = custom_field_mapping.get(field, field.replace('_', ' ').title())
        table_header += f"<th>{html.escape(field_name)}</th>"
    table_header += "</tr>"

    # Define column widths for fixed table layout
    colgroup = """
    <colgroup>
        <col style="width: 5%;">
        <col style="width: 15%;">
        <col style="width: 20%;">
        {col_widths}
    </colgroup>
    """.format(col_widths="".join(['<col style="width: 10%;">' for _ in fields]))

    table_rows = ""
    for i, issue in enumerate(issues, start=1):
        # Alternate row color: light gray for odd rows, white for even rows
        row_color = "#f2f2f2" if i % 2 != 0 else "#ffffff"
        table_row = f"<tr style='background-color:{row_color};'><td>{i}</td><td>{html.escape(issue['key'])}</td><td>{html.escape(issue['fields']['summary'])}</td>"

        for field in fields:
            value = issue['fields'].get(field, "")

            # Handle subtasks separately
            if field == 'subtasks':
                subtasks = issue['fields'].get('subtasks', [])
                subtask_keys = [subtask['key'] for subtask in subtasks]
                value = ', '.join(subtask_keys)

            # Handle customfield_10005 (Sprint Name)
            elif field == 'customfield_10005':
                if isinstance(value, list) and value:
                    value = value[0].split("name=")[-1].split(",")[0]  # Extract name from string
                else:
                    value = ""

            # Handle customfield_26424 (Requirement Status)
            elif field == 'customfield_26424':
                if isinstance(value, list) and value:
                    value = value[0].get('status', '')  # Extract status from dictionary
                else:
                    value = ""

            # If value is a dictionary with a 'name' key
            elif isinstance(value, dict) and 'name' in value:
                value = value['name']

            # If value is a list (but not handled earlier)
            elif isinstance(value, list):
                value = ', '.join(str(v['name'] if isinstance(v, dict) and 'name' in v else v) for v in value)

            # Replace None with an empty string to prevent merging issues
            if value is None:
                value = ""

            # Replace any "!" character in the value
            if isinstance(value, str):
                value = value.replace("!", "")

            # Escape the cell content to prevent HTML parsing issues
            table_row += f"<td>{html.escape(value)}</td>"

        table_row += "</tr>"
        table_rows += table_row

    # Combine header, rows, and column styles into the final table HTML
    html_table = f"""
    <table border='1' cellpadding='5' cellspacing='0' style='border-collapse: collapse; width: 100%; table-layout: fixed;'>
        {colgroup}
        {table_header}
        {table_rows}
    </table>
    """
    return html_table


table_row += f"<td>&nbsp;{escape(value)}&nbsp;</td>"
value = value.replace('\u200B', '')


########################################################################################################

def remove_special_characters(value):
    """Remove special characters from a value and attempt to convert back to original type."""
    # Convert value to string
    value_str = str(value).replace("!", "")
    
    # Attempt to convert back to float or int if possible
    if isinstance(value, (int, float)):
        try:
            value = float(value_str) if '.' in value_str else int(value_str)
        except ValueError:
            value = value_str  # Fallback to string if conversion fails
    else:
        value = value_str  # Use the cleaned string value

    return value

def generate_html_table(issues, fields):
    # Table header
    table_header = "<tr><th>Serial No</th><th>Story</th><th>Summary</th>"
    for field in fields:
        field_name = custom_field_mapping.get(field, field.replace('_', ' ').title())
        table_header += f"<th>{html.escape(field_name)}</th>"
    table_header += "</tr>"

    colgroup = """
    <colgroup>
        <col style="width: 5%;">
        <col style="width: 15%;">
        <col style="width: 20%;">
        {col_widths}
    </colgroup>
    """.format(col_widths="".join(['<col style="width: 10%;">' for _ in fields]))

    table_rows = ""
    for i, issue in enumerate(issues, start=1):
        row_color = "#f2f2f2" if i % 2 != 0 else "#ffffff"
        table_row = f"<tr style='background-color:{row_color};'><td>{i}</td><td>{html.escape(issue['key'])}</td><td>{html.escape(issue['fields']['summary'])}</td>"

        for field in fields:
            if field == 'subtasks':
                subtasks = issue['fields'].get('subtasks', [])
                subtask_keys = [subtask['key'] for subtask in subtasks]
                value = ', '.join(subtask_keys)
            else:
                value = issue['fields'].get(field, "")

            if field == 'customfield_10005' and isinstance(value, list) and value:
                value = value[0].split("name=")[-1].split(",")[0]

            elif field == 'customfield_26424' and isinstance(value, list) and value:
                value = value[0].get('status', '')

            elif isinstance(value, dict) and 'displayName' in value:
                value = value['displayName']
            elif isinstance(value, dict) and 'name' in value:
                value = value['name']
            elif isinstance(value, dict) and 'value' in value:
                value = value['value']
            elif isinstance(value, list):
                value = ', '.join(str(v['name'] if isinstance(v, dict) and 'name' in v else v) for v in value)

            if value is None or value == '':
                value = "Not available"

            # Remove special characters and attempt to convert back to original type
            value = remove_special_characters(value)

            table_row += f"<td>{html.escape(str(value))}</td>"

        table_row += "</tr>"
        table_rows += table_row

    html_table = f"""
    <table border='1' cellpadding='5' cellspacing='0' style='border-collapse: collapse; width: 100%; table-layout: fixed;'>
        {colgroup}
        {table_header}
        {table_rows}
    </table>
    """
    return html_table



### add comments


def generate_html_table(issues, fields):
    # Table header
    table_header = "<tr><th>Serial No</th><th>Story</th><th>Summary</th>"
    for field in fields:
        field_name = custom_field_mapping.get(field, field.replace('_', ' ').title())
        table_header += f"<th>{html.escape(field_name)}</th>"
    table_header += "<th>Comments</th></tr>"  # Add Comments column to the header

    # Define column widths for fixed table layout
    colgroup = """
    <colgroup>
        <col style="width: 5%;">
        <col style="width: 15%;">
        <col style="width: 20%;">
        {col_widths}
    </colgroup>
    """.format(col_widths="".join(['<col style="width: 10%;">' for _ in fields]) + '<col style="width: 10%;">')

    table_rows = ""
    for i, issue in enumerate(issues, start=1):
        row_color = "#f2f2f2" if i % 2 != 0 else "#ffffff"
        table_row = f"<tr style='background-color:{row_color};'><td>{i}</td><td>{html.escape(issue['key'])}</td><td>{html.escape(issue['fields']['summary'])}</td>"
        
        comment = "NA"  # Default value for the Comments column

        for field in fields:
            value = issue['fields'].get(field, "")
            if field == 'subtasks':
                subtasks = issue['fields'].get('subtasks', [])
                subtask_keys = [subtask['key'] for subtask in subtasks]
                value = ', '.join(subtask_keys)
            elif field == 'customfield_10005' and isinstance(value, list) and value:
                value = value[0].split("name=")[-1].split(",")[0]
            elif field == 'customfield_26424' and isinstance(value, list) and value:
                value = value[0].get('status', '')  # Extract status from dictionary
                requirement_status = value 
            elif isinstance(value, dict) and 'displayName' in value:
                value = value['displayName']
            elif isinstance(value, dict) and 'name' in value:
                value = value['name']
            elif isinstance(value, dict) and 'value' in value:
                value = value['value']
            elif isinstance(value, list):
                value = ', '.join(str(v['name'] if isinstance(v, dict) and 'name' in v else v) for v in value)
            
            if value is None or value == "":
                value = "Not Available"
            else:
                value = remove_special_characters(value)
            
            cell_content = html.escape(str(value))

            if field == "customfield_20627":  # Assuming this is the QA Assignee field
                if qa_assignee != "Not Available":
                    if qa_required != "Yes" and requirement_status != "OK":
                        table_row += f"<td style='background-color: yellow;'>{cell_content}</td>"
                        comment = "Yellow column"
                    elif qa_required == "Not Available" and requirement_status == "OK":
                        table_row += f"<td style='background-color: red;'>{cell_content}</td>"
                        comment = "Red column"
                    else:
                        if (qa_required == "Yes" and requirement_status != "OK") or (qa_required == "No" and requirement_status == "OK"):
                            table_row += f"<td style='background-color: red;'>{cell_content}</td>"
                            comment = "Red column"
                        else:
                            table_row += f"<td>{cell_content}</td>"
                elif qa_assignee == "Not Available":
                    if requirement_status == "OK" or qa_required == "Yes":
                        table_row += f"<td style='background-color: blue;'>{cell_content}</td>"
                        comment = "Blue column"
                    else:
                        table_row += f"<td>{cell_content}</td>"
            else:
                table_row += f"<td>{cell_content}</td>"

        table_row += f"<td>{html.escape(comment)}</td>"  # Add the Comments column value
        table_row += "</tr>"
        table_rows += table_row

    return f"<table>{colgroup}{table_header}{table_rows}</table>"


##333 rnhanced conditions

for field in fields:
    value = issue['fields'].get(field, "")

    if field == 'subtasks':
        # Handle subtasks as a list of keys
        subtasks = value if isinstance(value, list) else []
        subtask_keys = [subtask['key'] for subtask in subtasks if 'key' in subtask]
        value = ', '.join(subtask_keys)

    elif field == 'customfield_10005':
        # Handle customfield_10005 as a list, extract 'name' if present
        if isinstance(value, list) and value:
            # Ensure that the string split logic only occurs if it's a string
            if isinstance(value[0], str) and "name=" in value[0]:
                value = value[0].split("name=")[-1].split(",")[0]
            else:
                value = "Not Available"
        else:
            value = "Not Available"

    elif field == 'customfield_26424':
        # Handle customfield_26424 as a list of dictionaries, extract 'status'
        if isinstance(value, list) and value:
            status_dict = value[0] if isinstance(value[0], dict) else {}
            value = status_dict.get('status', "Not Available")
            requirement_status = value  # Preserved for use later in highlighting logic
        else:
            value = "Not Available"

    elif isinstance(value, dict):
        # Handle dictionary fields like 'displayName', 'name', or 'value'
        if 'displayName' in value:
            value = value['displayName']
        elif 'name' in value:
            value = value['name']
        elif 'value' in value:
            value = value['value']
        else:
            value = "Not Available"

    elif isinstance(value, list):
        # Handle list fields by extracting names or using string representations
        value = ', '.join(
            str(v['name'] if isinstance(v, dict) and 'name' in v else v) 
            for v in value
        ) if value else "Not Available"

    # Handle cases where value is None or empty
    if not value:
        value = "Not Available"

    cell_content = html.escape(str(value))
    # Process cell_content according to your existing logic


if field in ['subtasks', 'customfield_26027', 'customfield_26424']:
            table_header += f"<th style='background-color: #D3D3D3;'>{html.escape(field_name)}</th>"
        else:
            table_header += f"<th>{html.escape(field_name)}</th>"
    
    table_header += "</tr>"


table_style = """
    <style>
        table {
            border-collapse: separate;
            border-spacing: 0;
            width: 100%;
            font-family: Arial, sans-serif;
            border-radius: 10px; /* Rounded corners */
            overflow: hidden;
        }
        th, td {
            padding: 8px;
            text-align: left;
            border: 1px solid #ddd;
        }
        th {
            background-color: #f2f2f2;
            font-weight: bold;
        }
        tr:nth-child(even) {
            background-color: #f9f9f9;
        }
        tr:hover {
            background-color: #f1f1f1;
        }
    </style>
    """
return table_style + "<table>" + colgroup + table_header + table_rows + "</table>"

............................

    table_style = """
<style>
    table {
        border-collapse: separate;
        border-spacing: 0;
        width: 100%;
        font-family: Arial, sans-serif;
        border-radius: 10px; /* Rounded corners */
        overflow: hidden;
        table-layout: fixed; /* Fixed layout for consistent column widths */
    }
    th, td {
        padding: 8px;
        text-align: left;
        border: 1px solid #ddd;
        word-wrap: break-word; /* Ensures long content wraps within the cell */
        max-width: 150px; /* Max width for cells to prevent them from expanding too much */
    }
    th {
        background-color: #f2f2f2;
        font-weight: bold;
    }
    tr:nth-child(even) {
        background-color: #f9f9f9;
    }
    tr:hover {
        background-color: #f1f1f1;
    }
</style>
"""

s_column = Date.parse("MM/dd/yyyy HH.mm.ss", s_column).format("yyyy-MM-dd")

SELECT TO_CHAR(CAST('2024-07-19 00:00:00' AS TIMESTAMP), 'MM/DD/YYYY HH24.MI.SS') 
FROM your_table;


###################################333 acceptance criteria condition #############################


def generate_html_table(issues, fields):
    # Table header
    table_header = "<tr><th>Serial No</th><th>Story</th><th>Summary</th>"
    for field in fields:
        # Replace custom field IDs with user-friendly names if available
        field_name = custom_field_mapping.get(field, field.replace('_', ' ').title())
        table_header += f"<th>{html.escape(field_name)}</th>"
    table_header += "</tr>"

    # Define column widths for fixed table layout
    colgroup = """
    <colgroup>
        <col style="width: 5%;">
        <col style="width: 15%;">
        <col style="width: 20%;">
        {col_widths}
    </colgroup>
    """.format(col_widths="".join(['<col style="width: 10%;">' for _ in fields]))

    table_rows = ""
    for i, issue in enumerate(issues, start=1):
        # Alternate row color: light gray for odd rows, white for even rows
        row_color = "#f2f2f2" if i % 2 != 0 else "#ffffff"
        table_row = f"<tr style='background-color:{row_color};'><td>{i}</td><td>{html.escape(issue['key'])}</td><td>{html.escape(issue['fields']['summary'])}</td>"

        # Initialize the comments and other variables
        combined_comment = "No issues"
        acceptance_criteria_comment = ""
        qa_required_comment = ""

        # Extract acceptance criteria
        acceptance_criteria = issue['fields'].get('customfield_1110', '')
        acceptance_length = len(acceptance_criteria)

        for field in fields:
            value = issue['fields'].get(field, "")
            if field == 'subtasks':
                subtasks = issue['fields'].get('subtasks', [])
                subtask_keys = [subtask['key'] for subtask in subtasks]
                value = ', '.join(subtask_keys)
            elif field == 'customfield_10005' and isinstance(value, list) and value:
                value = value[0].split("name=")[-1].split(",")[0]
            elif field == 'customfield_26424' and isinstance(value, list) and value:
                value = value[0].get('status', '')
                requirement_status = value
            elif isinstance(value, dict):
                if 'displayName' in value:
                    value = value['displayName']
                elif 'name' in value:
                    value = value['name']
                elif 'value' in value:
                    value = value['value']
            elif isinstance(value, list):
                value = ', '.join(str(v['name'] if isinstance(v, dict) and 'name' in v else v) for v in value)

            if value is None or value == "":
                value = "Not Available"
            else:
                value = remove_special_characters(value)

            # Capture QA Required? for logic check and clean it up
            if field == 'customfield_26027':
                qa_required = str(value).replace("!", "").strip()

            elif field == 'customfield_17201':
                qa_assignee = str(value).replace("!", "").strip()

            # Escape the cell content to prevent HTML parsing issues
            cell_content = html.escape(str(value))

            # Apply the acceptance criteria rule and update the comment
            if field == "customfield_1110":  # Acceptance Criteria
                if acceptance_length == 0:
                    table_row += f"<td style='background-color: blue;'>{cell_content}</td>"
                    acceptance_criteria_comment = "less"
                elif acceptance_length < 30:
                    table_row += f"<td style='background-color: blue;'>{cell_content}</td>"
                    acceptance_criteria_comment = "more"
                else:
                    table_row += f"<td>{cell_content}</td>"

            # Apply the QA Required? and Requirement Status rules and update the comment
            elif field == "customfield_20627":
                if qa_assignee != "Not Available":
                    if qa_required != "Yes" and requirement_status != "OK":
                        table_row += f"<td style='background-color: yellow;'>{cell_content}</td>"
                        qa_required_comment = "Yellow column"
                    elif qa_required == "Not Available":
                        if requirement_status == "OK":
                            table_row += f"<td style='background-color: red;'>{cell_content}</td>"
                            qa_required_comment = "Red column"
                        else:
                            table_row += f"<td>{cell_content}</td>"
                    else:
                        if (qa_required == "Yes" and requirement_status != "OK") or (qa_required == "No" and requirement_status == "OK"):
                            table_row += f"<td style='background-color: blue;'>{cell_content}</td>"
                            qa_required_comment = "Blue column"
                        else:
                            table_row += f"<td>{cell_content}</td>"
                elif qa_assignee == "Not Available":
                    if requirement_status == "OK" or qa_required == "Yes":
                        table_row += f"<td style='background-color: blue;'>{cell_content}</td>"
                        qa_required_comment = "Blue column"
                    else:
                        table_row += f"<td>{cell_content}</td>"
                else:
                    table_row += f"<td>{cell_content}</td>"

        # Combine comments from both conditions
        if acceptance_criteria_comment and qa_required_comment:
            combined_comment = f"{acceptance_criteria_comment}, {qa_required_comment}"
        elif acceptance_criteria_comment:
            combined_comment = acceptance_criteria_comment
        elif qa_required_comment:
            combined_comment = qa_required_comment

        # Finalize the row and add the combined comment
        table_row += f"<td>{html.escape(combined_comment)}</td>"
        table_row += "</tr>"
        table_rows += table_row

    return f"<table>{colgroup}{table_header}{table_rows}</table>"


    #############33 modified

    def generate_html_table(issues, fields):
    # Table header
    table_header = "<tr><th>Serial No</th><th>Story</th><th>Summary</th>"
    for field in fields:
        # Replace custom field IDs with user-friendly names if available
        field_name = custom_field_mapping.get(field, field.replace('_', ' ').title())
        table_header += f"<th>{html.escape(field_name)}</th>"
    table_header += "</tr>"

    # Define column widths for fixed table layout
    colgroup = """
    <colgroup>
        <col style="width: 5%;">
        <col style="width: 15%;">
        <col style="width: 20%;">
        {col_widths}
    </colgroup>
    """.format(col_widths="".join(['<col style="width: 10%;">' for _ in fields]))

    table_rows = ""
    for i, issue in enumerate(issues, start=1):
        # Alternate row color: light gray for odd rows, white for even rows
        row_color = "#f2f2f2" if i % 2 != 0 else "#ffffff"
        table_row = f"<tr style='background-color:{row_color};'><td>{i}</td><td>{html.escape(issue['key'])}</td><td>{html.escape(issue['fields']['summary'])}</td>"
        
        # Initialize the comment variable
        acceptance_criteria_comment = ""
        qa_required_comment = ""
        
        issue_type = issue['fields'].get('issuetype', {}).get('name', "")
        
        for field in fields:
            value = issue['fields'].get(field, "")
            if field == 'subtasks':
                subtasks = issue['fields'].get('subtasks', [])
                subtask_keys = [subtask['key'] for subtask in subtasks]
                value = ', '.join(subtask_keys)
            # Handle customfield_10005 (Sprint Name)
            elif field == 'customfield_10005' and isinstance(value, list) and value:
                value = value[0].split("name=")[-1].split(",")[0]  # Extract name from string

            # Handle customfield_26424 (Status)
            elif field == 'customfield_26424' and isinstance(value, list) and value:
                value = value[0].get('status', '')  # Extract status from dictionary

            elif isinstance(value, dict) and 'displayName' in value:
                value = value['displayName']
            elif isinstance(value, dict) and 'name' in value:
                value = value['name']
            elif isinstance(value, dict) and 'value' in value:
                value = value['value']	
            elif isinstance(value, list):
                value = ', '.join(str(v['name'] if isinstance(v, dict) and 'name' in v else v) for v in value)

            # Replace None with an empty string to prevent merging issues
            if value is None or value == "":
                value = "Not Available"
            else:
                value = remove_special_characters(value)  # Removing special characters

            # Capture QA Required? for logic check and clean it up
            if field == 'customfield_26027':
                qa_required = str(value).replace("!", "").strip()  # Remove '!' and trim whitespace
            
            elif field == 'customfield_17201':
                qa_assignee = str(value).replace("!", "").strip()  # Remove '!' and trim whitespace

            # Escape the cell content to prevent HTML parsing issues
            cell_content = html.escape(str(value))

            # Apply the acceptance criteria rule only for User Story issue type
            if issue_type == "User Story" and field == "customfield_1110":  # Acceptance Criteria
                acceptance_length = len(value) if value else 0
                if acceptance_length == 0:
                    table_row += f"<td style='background-color: blue;'>{cell_content}</td>"
                    acceptance_criteria_comment = "less"
                elif acceptance_length < 30:
                    table_row += f"<td style='background-color: blue;'>{cell_content}</td>"
                    acceptance_criteria_comment = "more"
                else:
                    table_row += f"<td>{cell_content}</td>"

            # Apply the highlighting rules
            if field == "customfield_20627":
                if qa_assignee != "Not Available":
                    if qa_required != "Yes" and requirement_status != "OK":
                        table_row += f"<td style='background-color: yellow;'>{cell_content}</td>"
                        qa_required_comment = "Yellow column"
                    elif qa_required == "Not Available":
                        # Case where QA Required? is Not available
                        if requirement_status == "OK":
                            table_row += f"<td style='background-color: red;'>{cell_content}</td>"
                            qa_required_comment = "Red column"
                        else:
                            table_row += f"<td>{cell_content}</td>"
                    else:
                        if (qa_required == "Yes" and requirement_status != "OK") or (qa_required == "No" and requirement_status == "OK"):
                            table_row += f"<td style='background-color: blue;'>{cell_content}</td>"
                            qa_required_comment = "Blue column"
                        else:
                            table_row += f"<td>{cell_content}</td>"
                elif qa_assignee == "Not Available":
                    if requirement_status == "OK" or qa_required == "Yes":
                        table_row += f"<td style='background-color: blue;'>{cell_content}</td>"
                        qa_required_comment = "Blue column"
                    else:
                        table_row += f"<td>{cell_content}</td>"

            # Finalize the cell content
            table_row += f"<td>{html.escape(cell_content)}</td>"

        # Merge the comments if available
        final_comment = acceptance_criteria_comment
       


qa_assignee = issue['fields'].get("customfield_17201", "Not Available")
qa_required = issue['fields'].get("customfield_20627", "Not Available")

# Ensure qa_required is not None before calling get
if qa_required and isinstance(qa_required, dict):
    qa_name = qa_required.get('value', 'Not Available')
else:
    qa_name = 'Not Available'
requirement_status = issue['fields'].get("customfield_26424", "Not Available")

################# clickabale link

import requests
from requests_kerberos import HTTPKerberosAuth, REQUIRED
import concurrent.futures
import time
import conf  # Import the configuration file
import logging
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import html  # For escaping HTML content

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime%s') - %(levelname)s - %(message)s')

# Jira server URL
jira_url = 'https://your-company-jira.com'

# Kerberos authentication
kerberos_auth = HTTPKerberosAuth(mutual_authentication=REQUIRED)

def fetch_all_issues(jql_query):
    start_at = 0
    max_results = 1000
    all_issues = []
    retry_attempts = 3
    retry_delay = 5

    while True:
        params = {
            'jql': jql_query,
            'startAt': start_at,
            'maxResults': max_results
        }

        for attempt in range(retry_attempts):
            try:
                response = requests.get(f'{jira_url}/rest/api/2/search', params=params, auth=kerberos_auth)

                if response.status_code == 200:
                    issues = response.json()['issues']
                    all_issues.extend(issues)
                    if len(issues) < max_results:
                        return all_issues
                    start_at += max_results
                    break
                else:
                    logging.error(f"Failed to retrieve issues for JQL '{jql_query}': {response.status_code} - {response.text}")
                    if response.status_code == 401:  # Unauthorized
                        logging.warning("Authentication issue, retrying...")
                        time.sleep(retry_delay)
            except requests.exceptions.RequestException as e:
                logging.error(f"Request failed: {e}")
                time.sleep(retry_delay)

        if attempt == retry_attempts - 1:
            logging.error(f"Failed to retrieve issues for JQL '{jql_query}' after {retry_attempts} attempts.")
            break

    return all_issues


def generate_html_table(issues, fields):
    # Table header
    table_header = "<tr><th>Serial No</th><th>Story</th><th>Summary</th>"
    for field in fields:
        # Replace custom field IDs with user-friendly names if available
        field_name = custom_field_mapping.get(field, field.replace('_', ' ').title())
        table_header += f"<th>{html.escape(field_name)}</th>"
    table_header += "</tr>"

    # Define column widths for fixed table layout
    colgroup = """
    <colgroup>
        <col style="width: 5%;">
        <col style="width: 15%;">
        <col style="width: 20%;">
        {col_widths}
    </colgroup>
    """.format(col_widths="".join(['<col style="width: 10%;">' for _ in fields]))

    table_rows = ""
    for i, issue in enumerate(issues, start=1):
        # Alternate row color: light gray for odd rows, white for even rows
        row_color = "#f2f2f2" if i % 2 != 0 else "#ffffff"
        
        # Create the link for the issue key
        issue_key_link = f'<a href="{jira_url}/browse/{issue["key"]}" target="_blank">{html.escape(issue["key"])}</a>'

        table_row = f"<tr style='background-color:{row_color};'><td>{i}</td><td>{issue_key_link}</td><td>{html.escape(issue['fields']['summary'])}</td>"
        
        # Initialize comment variables
        acceptance_criteria_comment = ""
        qa_required_comment = ""
        requirement_status = ""
        qa_required = ""
        qa_assignee = ""

        # Iterate over each field to populate the table row
        for field in fields:
            value = issue['fields'].get(field, "")
            if field == 'subtasks':
                subtasks = issue['fields'].get('subtasks', [])
                subtask_keys = [subtask['key'] for subtask in subtasks]
                value = ', '.join(subtask_keys)
            elif field == 'customfield_10005' and isinstance(value, list) and value:
                value = value[0].split("name=")[-1].split(",")[0]  # Extract name from string
            elif field == 'customfield_26424' and isinstance(value, list) and value:
                value = value[0].get('status', '')  # Extract status from dictionary
                requirement_status = value
            elif isinstance(value, dict) and 'displayName' in value:
                value = value['displayName']
            elif isinstance(value, dict) and 'name' in value:
                value = value['name']
            elif isinstance(value, dict) and 'value' in value:
                value = value['value']
            elif isinstance(value, list):
                value = ', '.join(str(v['name'] if isinstance(v, dict) and 'name' in v else v) for v in value)

            if value is None or value == "":
                value = "Not Available"
            else:
                value = remove_special_characters(value)  # Removing special characters
            
            if field == 'customfield_26027':
                qa_required = str(value).replace("!", "").strip()  # Remove '!' and trim whitespace
            elif field == 'customfield_17201':
                qa_assignee = str(value).replace("!", "").strip()  # Remove '!' and trim whitespace

            cell_content = html.escape(str(value))

            # Apply the highlighting rules
            if issue_type == "User Story" and field == "customfield_1110":  # Acceptance Criteria
                acceptance_length = len(value) if value else 0
                if acceptance_length == 0:
                    table_row += f"<td style='background-color: blue;'>{cell_content}</td>"
                    acceptance_criteria_comment = "less"
                elif acceptance_length < 30:
                    table_row += f"<td style='background-color: blue;'>{cell_content}</td>"
                    acceptance_criteria_comment = "more"
                else:
                    table_row += f"<td>{cell_content}</td>"

            elif field == "customfield_20627":
                if qa_assignee != "Not Available":
                    if qa_required != "Yes" and requirement_status != "OK":
                        table_row += f"<td style='background-color: yellow;'>{cell_content}</td>"
                        qa_required_comment = "Yellow column"
                    elif qa_required == "Not Available":
                        if requirement_status == "OK":
                            table_row += f"<td style='background-color: red;'>{cell_content}</td>"
                            qa_required_comment = "Red column"
                        else:
                            table_row += f"<td>{cell_content}</td>"
                    else:
                        if (qa_required == "Yes" and requirement_status != "OK") or (qa_required == "No" and requirement_status == "OK"):
                            table_row += f"<td style='background-color: blue;'>{cell_content}</td>"
                            qa_required_comment = "Blue column"
                        else:
                            table_row += f"<td>{cell_content}</td>"
                elif qa_assignee == "Not Available":
                    if requirement_status == "OK" or qa_required == "Yes":
                        table_row += f"<td style='background-color: blue;'>{cell_content}</td>"
                        qa_required_comment = "Blue column"
                    else:
                        table_row += f"<td>{cell_content}</td>"
                else:
                    table_row += f"<td>{cell_content}</td>"

        if acceptance_criteria_comment and qa_required_comment:
            combined_comment = f"{acceptance_criteria_comment}, {qa_required_comment}"
        elif acceptance_criteria_comment:
            combined_comment = acceptance_criteria_comment
        elif qa_required_comment:
            combined_comment = qa_required_comment

        table_row += f"<td>{html.escape(combined_comment)}</td>"
        table_row += "</tr>"
        table_rows += table_row

    # Combine table header, colgroup, and table rows to complete the HTML table
    html_table = f"<table border='1' style='border-collapse: collapse;'>{colgroup}{table_header}{table_rows}</table>"
    return html_table


def process_issues(jql_query, all_email_content, fields):
    issues = fetch_all_issues(jql_query)
    if not issues:
        return

    component_name = jql_query.split('component = ')[1].split()[0].replace('"', '')
    email_content = f"<h2>Results for component: {component_name}</h2><br>"

    email_content += generate_html_table(issues, fields)

    all_email_content.append(email_content)

def send_email(subject, body):
    sender_email = "your_email@example.com"
    recipient_email = "recipient@example.com"

    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = recipient_email
    msg['Subject'] = subject

    msg.attach(MIMEText(body, 'html'))

    try:
        server = smtplib.SMTP('smtp.example.com', 587)
        server.starttls()
        server.login(sender_email, "your_password")
        server.sendmail(sender_email, recipient_email, msg.as_string())
        server.quit()
        logging.info(f"Email sent to {recipient_email}")
    except Exception as

  # Add the JQL query itself to the email content
    # Add the JQL query with blue color styling
    email_content += f"<p><strong>JQL:</strong> <span style='color:blue;'>{html.escape(jql_query)}</span></p><br>"


def process_issues(jql_query, all_email_content, fields):
    issues = fetch_all_issues(jql_query)
    component_name = jql_query.split('component = ')[1].split()[0].replace('"', '')
    
    email_content = f"<h2>Results for component: {component_name}</h2><br>"

    if issues:
        email_content += generate_html_table(issues, fields)
    else:
        email_content += "<p style='color:red;'>No data available for this JQL.</p>"

    all_email_content.append(email_content)



# subtask with status

for field in fields:
            value = issue['fields'].get(field, "")
            if field == 'subtasks':
                subtasks = issue['fields'].get('subtasks', [])
                subtask_values = [f"{subtask['key']} ({subtask['fields']['status']['name']})" for subtask in subtasks]
                value = ', '.join(subtask_values)



 elif field == 'issuelinks':
    issue_links = []
    relevant_statuses = []  # List to store statuses of relevant issue types
    
    # Check if 'issuelinks' field exists in the response
    if 'issuelinks' in issue.get('fields', {}):
        for link in issue['fields']['issuelinks']:
            if 'inwardIssue' in link:
                inward_issue = link['inwardIssue']
                link_key = inward_issue['key']
                link_status = inward_issue['fields']['status']['name']
                link_type = inward_issue['fields']['issuetype']['name']
                if link_type in ['Bug', 'Defect', 'Defect Sub-Task']:
                    issue_links.append(f"{link_key} ({link_status})")
                    relevant_statuses.append(link_status)  # Store status
            if 'outwardIssue' in link:
                outward_issue = link['outwardIssue']
                link_key = outward_issue['key']
                link_status = outward_issue['fields']['status']['name']
                link_type = outward_issue['fields']['issuetype']['name']
                if link_type in ['Bug', 'Defect', 'Defect Sub-Task']:
                    issue_links.append(f"{link_key} ({link_status})")
                    relevant_statuses.append(link_status)  # Store status
    
    # Handle the case where 'issuelinks' doesn't exist or no relevant issue types found
    if not issue_links:
        value = "No linked defects available"
    else:
        value = ', '.join(issue_links)

    # You can now use the `relevant_statuses` list for future purposes
    # For example, you might store it in a dictionary or process it later


def check_status_list(status_list):
    allowed_statuses = {"CLOSED", "CANCELLED"}  # Set of allowed statuses (case-insensitive)

    # If the status list is empty, do nothing (return True by default)
    if not status_list:
        return True

    # Convert each status to uppercase and check if it's in the allowed set
    for status in status_list:
        if status.upper() not in allowed_statuses:
            return False

# Example usage
abc = ['CLOSED', 'COOL']
result = check_status_list(abc)
print(result)  # Output will be False


for field in fields:
            value = issue['fields'].get(field, "")
            if field == 'subtasks':
                subtasks = issue['fields'].get('subtasks', [])
                if subtasks:
                    subtask_values = [f"{index + 1}. {subtask['key']} ({subtask['fields']['status']['name']})"
                                      for index, subtask in enumerate(subtasks)]
                    value = '<br>'.join(subtask_values)  # Join with line breaks for numbered list format




# mulitplessssssssssss


import csv
import sys
import logging

def process_all_jql_queries(jql_queries, fields):
    all_email_content = []

    for jql_query in jql_queries:
        logging.info(f"Processing JQL query: {jql_query}")
        process_issues(jql_query, all_email_content, fields)

    return ''.join(all_email_content)

def main():
    fields = [
        'issuetype',
        'priority',
        'versions',
        'components',
        'labels',
        'status',
        'resolution',
        'fixVersions',
        'customfield_10021',  # QA Required? (custom field)
        'customfield_10020',  # Sprint (custom field)
        'customfield_10002',  # Story Points (custom field)
    ]

    if len(sys.argv) < 3:
        logging.error("Insufficient arguments provided. Please provide 'squad' or 'feed' as the first argument.")
        return

    # Condition for 'squad'
    if sys.argv[1] == "squad":
        squad_names = []
        squad_names.append(sys.argv[2])  # Get squad names from argument
        squad_list = ', '.join(squad_names)
        jql_queries = [
            f'{sys.argv[3]} and type in ("User Story", "Sub-task") and component = "{component}"'
            for component in squad_names
        ]
        subject = f"JIRA Compliance report for {sys.argv[1]} {squad_list} for {sys.argv[3]}"
        recipient = "abc@gmail.com"

        # Process JQL queries
        email_body = process_all_jql_queries(jql_queries, fields)

        # Send the email
        send_email(subject, email_body, recipient)
        logging.info(f"Email sent to {recipient}")

    # Condition for 'feed'
    elif sys.argv[1] == "feed":
        csv_file = sys.argv[2]  # CSV filename
        try:
            with open(csv_file, 'r') as file:
                csv_reader = csv.DictReader(file)

                for row in csv_reader:
                    squad = row['squad']
                    squad_name = row['squad_name']
                    condition = row['condition']
                    emails = row['email'].split(',')

                    # Construct JQL query using squad name and condition from CSV
                    jql_query = [f"squad = {squad_name} AND {condition}"]

                    # Process JQL query and fetch the results
                    email_body = process_all_jql_queries(jql_query, fields)

                    # Send the emails to all recipients listed in the CSV
                    email_subject = f"Jira report for {squad_name}"
                    for recipient_email in emails:
                        send_email(email_subject, email_body, recipient_email)
                        logging.info(f"Sent email to {recipient_email}")

        except FileNotFoundError:
            logging.error(f"CSV file '{csv_file}' not found.")
        except KeyError as e:
            logging.error(f"Missing column in CSV: {e}")
    else:
        logging.error("Invalid argument passed. First argument should be either 'squad' or 'feed'.")

if __name__ == "__main__":
    main()


squad|squad_name|condition|email
squad|abc-drm|fixVersion=35.45.44|abc@gmail.com, fpay@gmail.com
squad|dnv sfsf|createddate > startofyear()|vnn@fkfk.com



elif sys.argv[1] == "feed":
    csv_file = sys.argv[2]  # CSV filename
    required_columns = ['squad', 'squad_name', 'condition', 'email']

    try:
        with open(csv_file, 'r') as file:
            csv_reader = csv.DictReader(file)

            # Check if all required columns are present
            missing_columns = [col for col in required_columns if col not in csv_reader.fieldnames]
            if missing_columns:
                logging.error(f"Missing required columns in CSV: {', '.join(missing_columns)}")
                return  # Exit if any columns are missing

            for row in csv_reader:
                try:
                    squad = row['squad']
                    squad_name = row['squad_name']
                    condition = row['condition']
                    emails = row['email'].split(',')




elif sys.argv[1] == "feed":
    csv_file = sys.argv[2]  # CSV filename
    required_columns = ['squad', 'squad_name', 'condition', 'email']

    try:
        with open(csv_file, 'r') as file:
            csv_reader = csv.DictReader(file, delimiter='|')

            # Check if all required columns are present
            missing_columns = [col for col in required_columns if col not in csv_reader.fieldnames]
            if missing_columns:
                logging.error(f"Missing required columns in CSV: {', '.join(missing_columns)}")
                sys.exit(1)

            for row in csv_reader:
                try:
                    squad = row['squad'].strip()
                    squad_names = row['squad_name'].split()  # Split based on spaces or other criteria
                    squad_list = ', '.join(squad_names)
                    condition = row['condition']
                    emails = row['email'].split(',')
                    
                    # Loop over each squad_name and generate individual JQL query
                    for squad_name in squad_names:
                        jql_query = f'{condition} and component="{squad_name.strip()}"'
                        print(f'Generated JQL: {jql_query}')  # Debug print to verify the JQL query

                        # Fetch issues and process them
                        all_email_content = []
                        process_issues(jql_query, all_email_content, fields=['summary', 'status'])  # Customize fields as needed

                        # Compose and send email for this squad
                        subject = f"Jira compliance for squad {squad_name.strip()} for {condition}"
                        recipient = ', '.join(emails)
                        body = '\n'.join(all_email_content)

                        send_email(subject, body)
                except Exception as e:
                    logging.error(f"Error processing row: {e}")
    except FileNotFoundError:
        logging.error(f"CSV file {csv_file} not found.")

















elif sys.argv[1] == "feed":
    csv_file = sys.argv[2]  # CSV filename
    required_columns = ['squad', 'squad_name', 'condition', 'email']

    try:
        with open(csv_file, 'r') as file:
            csv_reader = csv.DictReader(file, delimiter='|')

            # Check if all required columns are present
            missing_columns = [col for col in required_columns if col not in csv_reader.fieldnames]
            if missing_columns:
                logging.error(f"Missing required columns in CSV: {', '.join(missing_columns)}")
                sys.exit(1)

            for row in csv_reader:
                try:
                    squad = row['squad'].strip()
                    squad_names = row['squad_name'].split()  # Split based on spaces or other criteria
                    squad_list = ', '.join(squad_names)
                    condition = row['condition']
                    emails = row['email'].split(',')
                    
                    # Loop over each squad_name and generate individual JQL query
                    for squad_name in squad_names:
                        jql_query = f'{condition} and component="{squad_name.strip()}"'
                        print(f'Generated JQL: {jql_query}')  # Debug print to verify the JQL query

                        # Fetch issues and process them
                        all_email_content = []
                        process_issues(jql_query, all_email_content, fields=['summary', 'status'])  # Customize fields as needed

                        # Compose and send email for this squad
                        subject = f"Jira compliance for squad {squad_name.strip()} for {condition}"
                        recipient = ', '.join(emails)
                        body = '\n'.join(all_email_content)

                        send_email(subject, body)
                except Exception as e:
                    logging.error(f"Error processing row: {e}")
    except FileNotFoundError:
        logging.error(f"CSV file {csv_file} not found.")



HTML to excel

import pandas as pd

# Read the HTML file
html_file = '/path_to_your_html_file/report_DRM_Lending.html'

# Parse the table(s) in the HTML file
tables = pd.read_html(html_file)

# Convert the first table to a DataFrame
df = tables[0]

# Save the DataFrame to an Excel file
output_excel_file = '/path_to_save_excel/report_DRM_Lending.xlsx'
df.to_excel(output_excel_file, index=False)

print(f"Excel file saved to {output_excel_file}")




from bs4 import BeautifulSoup
import openpyxl

# Load the HTML content from file
with open('/path_to_your_html_file/report_DRM_Lending.html', 'r') as f:
    html_content = f.read()

# Parse the HTML content
soup = BeautifulSoup(html_content, 'html.parser')

# Create a new Excel workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "DRM Lending Report"

# Find the table in the HTML (assuming there's one table, adjust if there are multiple)
table = soup.find('table')

# Extract the table headers
headers = [th.get_text(strip=True) for th in table.find_all('th')]
ws.append(headers)  # Write headers to the first row

# Extract the table rows and write them to the Excel file
for row in table.find_all('tr')[1:]:  # Skip the header row
    cells = [td.get_text(strip=True) for td in row.find_all('td')]
    ws.append(cells)  # Write the row data to Excel

# Save the Excel file
output_excel_file = '/path_to_save_excel/report_DRM_Lending.xlsx'
wb.save(output_excel_file)

print(f"Excel file saved to {output_excel_file}")



{
    "workbench.colorCustomizations": {
        "editor.background": "#f5f5f7",                // Soft light background like Windows 11 light mode
        "editor.foreground": "#1f1f1f",                // Dark text for readability
        "editorLineNumber.foreground": "#8a8a8a",      // Muted gray for line numbers
        "editor.selectionBackground": "#a2c9ff",       // Light blue for selected text, like Windows highlight
        "editor.selectionHighlightBackground": "#cfe5ff", // Soft blue for highlighted selections
        "editorCursor.foreground": "#0078d4",          // Windows blue for the cursor
        "editorBracketMatch.background": "#e0e0e0",    // Subtle gray for matching brackets
        "editorBracketMatch.border": "#0078d4",        // Windows blue border around matching brackets
        "editorWhitespace.foreground": "#e0e0e0",      // Light gray for whitespace characters

        "sideBar.background": "#f3f3f5",               // Light gray for the sidebar
        "sideBar.foreground": "#333333",               // Dark gray for sidebar text
        "sideBarSectionHeader.background": "#0078d4",  // Windows blue for the sidebar headers
        "sideBarSectionHeader.foreground": "#ffffff",  // White text for the sidebar header

        "activityBar.background": "#0078d4",           // Windows 11 blue for the activity bar
        "activityBar.foreground": "#ffffff",           // White icons for the activity bar
        "activityBarBadge.background": "#ff4500",      // Bright orange badges (like Windows notifications)
        "activityBarBadge.foreground": "#ffffff",      // White text for badges

        "statusBar.background": "#0078d4",             // Windows blue for the status bar
        "statusBar.foreground": "#ffffff",             // White text for the status bar
        "statusBar.noFolderBackground": "#404040",     // Dark gray for no folder status
        "statusBar.debuggingBackground": "#005a9e",    // Darker blue for debugging mode

        "titleBar.activeBackground": "#f3f3f5",        // Light gray title bar (like Windows title bar)
        "titleBar.activeForeground": "#333333",        // Dark gray for title bar text
        "titleBar.inactiveBackground": "#e0e0e0",      // Slightly darker for inactive windows
        "titleBar.inactiveForeground": "#8a8a8a",      // Muted gray for inactive title text

        "button.background": "#0078d4",                // Windows blue for buttons
        "button.foreground": "#ffffff",                // White button text
        "button.hoverBackground": "#005a9e",           // Darker blue for hover state

        "dropdown.background": "#ffffff",              // White background for dropdowns
        "dropdown.foreground": "#1f1f1f",              // Dark text for dropdown options
        "dropdown.border": "#e0e0e0",                  // Light gray border for dropdowns

        "input.background": "#ffffff",                 // White input fields
        "input.foreground": "#1f1f1f",                 // Dark text in input fields
        "input.border": "#e0e0e0",                     // Light gray borders for inputs
        "input.placeholderForeground": "#8a8a8a",      // Muted gray for placeholder text

        "tab.activeBackground": "#ffffff",             // Active tab with white background
        "tab.activeForeground": "#0078d4",             // Windows blue text for active tab
        "tab.inactiveBackground": "#f3f3f5",           // Light gray background for inactive tabs
        "tab.inactiveForeground": "#8a8a8a",           // Muted gray for inactive tab text
        "tab.border": "#e0e0e0",                       // Light gray border for tabs
        "tab.hoverBackground": "#0078d4",              // Blue hover effect on tabs

        "scrollbarSlider.background": "#d0d0d0",       // Subtle gray scrollbar
        "scrollbarSlider.hoverBackground": "#a0a0a0",  // Darker gray for hover state on scrollbar
        "scrollbarSlider.activeBackground": "#808080", // Dark gray for active scrollbar

        "panel.background": "#f5f5f7",                 // Light background for panel
        "panel.border": "#e0e0e0",                     // Light gray border for panel

        "terminal.background": "#1f1f1f",              // Dark terminal background
        "terminal.foreground": "#f5f5f7",              // Light text for terminal
        "terminalCursor.background": "#f5f5f7",        // White cursor in terminal
        "terminalCursor.foreground": "#0078d4",        // Windows blue for terminal cursor
        "terminal.selectionBackground": "#a2c9ff",     // Light blue for selected text in terminal

        "badge.background": "#ff4500",                 // Orange badge like Windows notification
        "badge.foreground": "#ffffff"                  // White text in badges
    },
    "editor.tokenColorCustomizations": {
        "comments": "#6a6a6a",                          // Muted gray for comments
        "keywords": "#0078d4",                          // Windows blue for keywords
        "functions": "#d83b01",                         // Orange for functions (like Windows app icons)
        "strings": "#228822",                           // Soft green for strings
        "variables": "#0078d4",                         // Blue for variables
        "types": "#ff8c00"                              // Orange for types (similar to Windows 11 icons)
    }
}


"workbench.colorCustomizations": {
  "activityBar.foreground": "#ffcc00", // Color of the activity bar icons (default state)
  "activityBar.activeBorder": "#00ff00", // Border color when the icon is active
  "activityBar.inactiveForeground": "#999999", // Color of inactive icons
  "activityBarBadge.background": "#ff0000", // Background color of the badges on the activity bar
  "activityBarBadge.foreground": "#ffffff"  // Foreground (text) color of the badges
}
"list.deemphasizedForeground": "#ffcc00" 


"workbench.colorCustomizations": {
  "tab.activeForeground": "#ffffff", // Text color for the active tab
  "tab.inactiveForeground": "#999999" // Text color for inactive tabs
}

"workbench.colorCustomizations": {
  "tab.activeBackground": "#1e1e1e",  // Background color of the active tab
  "tab.activeBorder": "#ffcc00",      // Border color of the active tab
  "tab.unfocusedActiveForeground": "#cccccc"  // Text color for the active tab when the editor is not focused
}


"workbench.colorCustomizations": {
  "list.focusBackground": "#ffcc00",   // Background color for focused item (when the sidebar is active and selected)
  "list.activeSelectionBackground": "#00ffcc",  // Background color of actively selected item
  "list.activeSelectionForeground": "#ffffff",  // Text color of actively selected item
  "list.inactiveSelectionBackground": "#999999",  // Background color of selected item when the sidebar is not focused
  "list.inactiveSelectionForeground": "#ffffff"   // Text color of selected item when the sidebar is not focused
}



# Initialize the counts
    total_records = len(issues)
    records_with_issues = 0
    
    table_rows = ""

# If there are comments other than "No Issues", consider it as having an issue
        if combined_comment != "":
            comment = combined_comment
            records_with_issues += 1  # Increment the counter for records with issues

        # Finalize the row and add the combined comment


 # Print total number of records and records with issues before returning the table HTML
    print(f"Total number of records: {total_records}")
    print(f"Total number of records with issues: {records_with_issues}")




# Parse the HTML using BeautifulSoup
soup = BeautifulSoup(html, 'html.parser')

# Find all <span> elements and extract their text
span_values = [span.get_text() for span in soup.find_all('span')]

# Join the values with commas
result = ', '.join(span_values)

print(result)

import re

# Open the HTML file
with open('your_file.html', 'r', encoding='utf-8') as file:
    html_content = file.read()

# Extract content between the phrase and the <h2> tag
match = re.search(r'check the boxes to filter the results(.*?)<h2>Results</h2>', html_content, re.DOTALL)

# If the section is found, extract <span> tags in that section
if match:
    filtered_html = match.group(1)
    span_values = re.findall(r'<span.*?>(.*?)</span>', filtered_html)
    
    # Join the values with commas
    result = ', '.join(span_values)
    print(result)
else:
    print("Section not found")





import pandas as pd

# Load JSON data (from a file or directly as a dictionary)
json_data = {
    "employees": [
        {"name": "John Doe", "age": 30, "department": "Engineering"},
        {"name": "Jane Smith", "age": 25, "department": "Marketing"},
        {"name": "Sam Brown", "age": 40, "department": "Sales"}
    ]
}

# Convert JSON to DataFrame
df = pd.DataFrame(json_data['employees'])

# Save DataFrame to Excel file
excel_file = 'output.xlsx'
df.to_excel(excel_file, index=False)

print(f"Data has been written to {excel_file}")




def dunc():
    # Define your Jira base URL
    jira_base_url = "https://ajira.com/browse"

    # Initialize variables
    tot_cnt_stories = 0
    tot_qa_required_stories = 0
    tot_cnt_test_plan = 0
    cnt_functional_test_plan = 0
    cnt_auto_in_functional = 0
    cnt_functional_test_cases = 0
    cnt_regression_test_plan = 0
    cnt_auto_in_regression = 0
    cnt_regression_test_cases = 0

    # Sets to capture distinct components and customfield_17201 values
    distinct_components = set()
    distinct_customfield_values = set()

    # Defect-related variables
    total_defects = 0
    closed_defects = 0
    open_defects = 0
    open_defect_details = []

    for index, row in df_data.iterrows():
        try:
            issue_type = row['fields.issuetype.name']
            issue_qa_required = row.get('fields.customfield_17201.value', "No")
            issue_key = row['key']
            status = row['fields.status.name']
            summary = row.get('fields.summary', '')

            # Collect distinct component values
            components = row.get('fields.components', [])
            component_names = [component.get('name') for component in components]
            for component_name in component_names:
                distinct_components.add(component_name)

            # Collect distinct customfield_17201 values
            customfield_value = row.get('fields.customfield_17201.value')
            if customfield_value:
                distinct_customfield_values.add(customfield_value)

            # Defect handling
            if issue_type in ["Defect Sub-Task", "Bug", "Dev_Bug"]:
                total_defects += 1
                if status == "Closed":
                    closed_defects += 1
                else:
                    open_defects += 1
                    # Create a clickable link for the issue key
                    issue_key_link = f'<a href="{jira_base_url}/{issue_key}">{issue_key}</a>'
                    open_defect_details.append(f"{issue_key_link} - {summary} ({', '.join(component_names)})")

        except Exception as e:
            return {"error": True, "error_info": "Jira item not found!"}
                    
        if issue_type == "User Story":
            tot_cnt_stories += 1
            if issue_qa_required == "Yes":
                tot_qa_required_stories += 1
        elif issue_type == "Test Plan":
            testPlanID = issue_key
            tot_cnt_test_plan += 1
            testPlanType = ''
            if "fields.summary" in df_data.columns:
                summary = row['fields.summary'].split(" | ")
                if len(summary) > 2:
                    testPlanType = summary[1]  # get test plan type from summary
            test_case_cnts = get_test_cases_status(testPlanID)
            if testPlanType == 'Functional':
                cnt_functional_test_plan += 1
                cnt_auto_in_functional += test_case_cnts['cnt_auto']
                cnt_functional_test_cases += test_case_cnts['cnt_auto'] + test_case_cnts['cnt_manual']
            elif testPlanType == 'Regression':
                cnt_regression_test_plan += 1
                cnt_auto_in_regression += test_case_cnts['cnt_auto']
                cnt_regression_test_cases += test_case_cnts['cnt_auto'] + test_case_cnts['cnt_manual']

    # Prepare defect information in a single string
    defect_info = f"Total defects: {total_defects} Closed defects: {closed_defects} Open defects: {open_defects}"
    if open_defects > 0:
        defect_info += "\n\nOpen defects details:\n" + "\n".join(open_defect_details)

    # Print or send defect information (if sending via email, ensure it's in HTML format)
    print(defect_info)

    # Print distinct components and customfield_17201 values
    print("Distinct Components in scope:", ', '.join(distinct_components))
    print("Distinct customfield_17201 values in scope:", ', '.join(distinct_customfield_values))






#### user.py

import pandas as pd
from openpyxl import load_workbook

# Load the Excel file
file_path = 'jan_aug_2024_2024-10-11.xlsx'
excel_data = pd.read_excel(file_path, sheet_name=0)

# Define the statuses that map to each category
automated_status = ['Automated']
backlog_status = ['Partial automated', 'Not automated', 'Automatable but not automated']
manual_status = ['Not feasible not automated']

# Create a summary DataFrame
summary_df = (
    excel_data
    .groupby(['Name', 'Fleet', 'Overall Automation Status'])
    .size()
    .reset_index(name='Count')
    .pivot_table(
        index=['Name', 'Fleet'],
        columns='Overall Automation Status',
        values='Count',
        fill_value=0
    )
    .reset_index()
)

# Reindex to add missing columns with a default value of 0
all_statuses = automated_status + backlog_status + manual_status
summary_df = summary_df.reindex(columns=['Name', 'Fleet'] + all_statuses, fill_value=0)

# Calculate totals and assign to respective columns
summary_df['Automated'] = summary_df[automated_status].sum(axis=1)
summary_df['Backlog'] = summary_df[backlog_status].sum(axis=1)
summary_df['Manual'] = summary_df[manual_status].sum(axis=1)
summary_df['Total'] = summary_df[['Automated', 'Backlog', 'Manual']].sum(axis=1)

# Keep only the required columns
final_summary_df = summary_df[['Name', 'Fleet', 'Automated', 'Backlog', 'Manual', 'Total']]

# Append the new sheet to the workbook
with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    final_summary_df.to_excel(writer, sheet_name='Summary', index=False)

print("Data has been successfully written to the new sheet in the workbook.")


# Successfully installed openpyxl-3.1.5 pandas-2.2.3

## fleet.py

import pandas as pd
from openpyxl import load_workbook

# Load the Excel file
file_path = 'jan_aug_2024_2024-10-11.xlsx'
excel_data = pd.read_excel(file_path, sheet_name=0)

# Define the statuses that map to each category
automated_status = ['Automated']
backlog_status = ['Partial automated', 'Not automated', 'Automatable but not automated']
manual_status = ['Not feasible not automated']

# Create a summary DataFrame grouped by 'Fleet'
summary_df = (
    excel_data
    .groupby(['Fleet', 'Overall Automation Status'])
    .size()
    .reset_index(name='Count')
    .pivot_table(
        index='Fleet',
        columns='Overall Automation Status',
        values='Count',
        fill_value=0
    )
    .reset_index()
)

# Reindex to add missing columns with a default value of 0
all_statuses = automated_status + backlog_status + manual_status
summary_df = summary_df.reindex(columns=['Fleet'] + all_statuses, fill_value=0)

# Calculate totals and assign to respective columns
summary_df['Automated'] = summary_df[automated_status].sum(axis=1)
summary_df['Backlog'] = summary_df[backlog_status].sum(axis=1)
summary_df['Manual'] = summary_df[manual_status].sum(axis=1)
summary_df['Total'] = summary_df[['Automated', 'Backlog', 'Manual']].sum(axis=1)

# Calculate the % Automated column
summary_df['% Automated'] = (summary_df['Automated'] / summary_df['Total']) * 100

# Keep only the required columns
final_summary_df = summary_df[['Name', 'Squad', 'Automated', 'Backlog', 'Manual', 'Total', '% Automated']]

# Append the new sheet to the workbook
with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    final_summary_df.to_excel(writer, sheet_name='By User', index=False)

print("Data has been successfully written to the new sheet in the workbook.")



########### j2e.py

import pandas as pd
import datetime
import json
import openpyxl  # Explicitly importing openpyxl

# Define file paths
input_file = 'back.json'

# Function to convert JSON keys to more readable column names
def convert_key_to_column_name(key):
    return key.replace('_', ' ').title()

# Generate today's date for the output Excel file name
today_date = datetime.datetime.now().strftime('%Y-%m-%d')
output_file = f'jan_aug_2024_{today_date}.xlsx'

# Load JSON data in a memory-efficient manner
def read_json_in_chunks(file_path, chunk_size=1000):
    with open(file_path, 'r') as file:
        data = json.load(file)
        for i in range(0, len(data), chunk_size):
            yield data[i:i+chunk_size]

# Write data to Excel in a memory-efficient way
with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
    first_chunk = True
    for chunk in read_json_in_chunks(input_file):
        # Convert chunk to DataFrame
        chunk_df = pd.DataFrame(chunk)
        # Rename columns dynamically
        chunk_df.columns = [convert_key_to_column_name(col) for col in chunk_df.columns]
        # Append to Excel sheet
        chunk_df.to_excel(writer, index=False, sheet_name='Data', startrow=0 if first_chunk else writer.sheets['Data'].max_row, header=first_chunk)
        first_chunk = False  # Disable header after the first chunk

print(f'Data has been successfully written to {output_file}')


# New requirement: Fetch subtask summary and check for 'risk type' keyword (case-insensitive)
subtask_with_risk_type = any('risk type' in subtask['fields'].get('summary', '').lower() for subtask in subtasks)



# Replace the string "null" with None (actual null) in DataFrame1
df1 = df1.withColumn("column_name", when(col("column_name") == "null", None).otherwise(col("column_name")))

# Replace blank strings with None in DataFrame2
df2 = df2.withColumn("column_name", when(col("column_name") == "", None).otherwise(col("column_name")))







#### GROUPING....................


def dunc():
    # Define your Jira base URL
    jira_base_url = "https://ajira.com/browse"

    # Initialize variables
    tot_cnt_stories = 0
    tot_qa_required_stories = 0
    tot_cnt_test_plan = 0
    cnt_functional_test_plan = 0
    cnt_auto_in_functional = 0
    cnt_functional_test_cases = 0
    cnt_regression_test_plan = 0
    cnt_auto_in_regression = 0
    cnt_regression_test_cases = 0

    # Sets to capture distinct components and customfield_17201 values
    distinct_components = set()
    distinct_customfield_values = set()

    # Defect-related variables
    total_defects = 0
    closed_defects = 0
    open_defects = 0
    open_defects_by_component = {}

    for index, row in df_data.iterrows():
        try:
            issue_type = row['fields.issuetype.name']
            issue_qa_required = row.get('fields.customfield_17201.value', "No")
            issue_key = row['key']
            status = row['fields.status.name']
            summary = row.get('fields.summary', '')

            # Collect distinct component values
            components = row.get('fields.components', [])
            component_names = [component.get('name') for component in components]

            for component_name in component_names:
                distinct_components.add(component_name)

            # Collect distinct customfield_17201 values
            customfield_value = row.get('fields.customfield_17201.value')
            if customfield_value:
                distinct_customfield_values.add(customfield_value)

            # Defect handling
            if issue_type in ["Defect Sub-Task", "Bug", "Dev_Bug"]:
                total_defects += 1
                if status == "Closed":
                    closed_defects += 1
                else:
                    open_defects += 1
                    # Create a clickable link for the issue key
                    issue_key_link = f'<a href="{jira_base_url}/{issue_key}">{issue_key}</a>'
                    defect_entry = f"{issue_key_link} - {summary}"

                    # Group open defects by component
                    for component_name in component_names:
                        if component_name not in open_defects_by_component:
                            open_defects_by_component[component_name] = []
                        open_defects_by_component[component_name].append(defect_entry)

        except Exception as e:
            return {"error": True, "error_info": "Jira item not found!"}

        if issue_type == "User Story":
            tot_cnt_stories += 1
            if issue_qa_required == "Yes":
                tot_qa_required_stories += 1
        elif issue_type == "Test Plan":
            testPlanID = issue_key
            tot_cnt_test_plan += 1
            testPlanType = ''
            if "fields.summary" in df_data.columns:
                summary = row['fields.summary'].split(" | ")
                if len(summary) > 2:
                    testPlanType = summary[1]  # get test plan type from summary
            test_case_cnts = get_test_cases_status(testPlanID)
            if testPlanType == 'Functional':
                cnt_functional_test_plan += 1
                cnt_auto_in_functional += test_case_cnts['cnt_auto']
                cnt_functional_test_cases += test_case_cnts['cnt_auto'] + test_case_cnts['cnt_manual']
            elif testPlanType == 'Regression':
                cnt_regression_test_plan += 1
                cnt_auto_in_regression += test_case_cnts['cnt_auto']
                cnt_regression_test_cases += test_case_cnts['cnt_auto'] + test_case_cnts['cnt_manual']

    # Prepare defect information by component
    defect_info = f"Total defects: {total_defects} Closed defects: {closed_defects} Open defects: {open_defects}"
    
    if open_defects > 0:
        defect_info += "\n\nOpen defects details:\n"
        for component, defects in open_defects_by_component.items():
            defect_info += f"\n{component}\n"
            defect_info += "\n".join(defects)

    # Print or send defect information (if sending via email, ensure it's in HTML format)
    print(defect_info)




Executive Summary:
The MLOps QA team built a tool for real-time feature monitoring, cutting manual efforts by 70%, detecting data drifts, and enhancing accuracy and scalability.






def dunc():
    # Define your Jira base URL
    jira_base_url = "https://ajira.com/browse"

    # Initialize variables
    tot_cnt_stories = 0
    tot_qa_required_stories = 0
    tot_cnt_test_plan = 0
    cnt_functional_test_plan = 0
    cnt_auto_in_functional = 0
    cnt_functional_test_cases = 0
    cnt_regression_test_plan = 0
    cnt_auto_in_regression = 0
    cnt_regression_test_cases = 0

    # Sets to capture distinct components and customfield_17201 values
    distinct_components = set()
    distinct_customfield_values = set()

    # Defect-related variables
    total_defects = 0
    closed_defects = 0
    open_defects = 0
    
    # Dictionary to hold open defects by component
    open_defect_details_by_component = {}

    for index, row in df_data.iterrows():
        try:
            issue_type = row['fields.issuetype.name']
            issue_qa_required = row.get('fields.customfield_17201.value', "No")
            issue_key = row['key']
            status = row['fields.status.name']
            summary = row.get('fields.summary', '')

            # Collect distinct component values
            components = row.get('fields.components', [])
            component_names = [component.get('name') for component in components]
            for component_name in component_names:
                distinct_components.add(component_name)

            # Collect distinct customfield_17201 values
            customfield_value = row.get('fields.customfield_17201.value')
            if customfield_value:
                distinct_customfield_values.add(customfield_value)

            # Defect handling
            if issue_type in ["Defect Sub-Task", "Bug", "Dev_Bug"]:
                total_defects += 1
                if status == "Closed":
                    closed_defects += 1
                else:
                    open_defects += 1
                    
                    # Create a clickable link for the issue key
                    issue_key_link = f'<a href="{jira_base_url}/{issue_key}">{issue_key}</a>'
                    
                    # Append defect details to the corresponding component in the dictionary
                    for component_name in component_names:
                        if component_name not in open_defect_details_by_component:
                            open_defect_details_by_component[component_name] = []
                        open_defect_details_by_component[component_name].append(f"{issue_key_link} - {summary}")

        except Exception as e:
            return {"error": True, "error_info": "Jira item not found!"}
                    
        if issue_type == "User Story":
            tot_cnt_stories += 1
            if issue_qa_required == "Yes":
                tot_qa_required_stories += 1
        elif issue_type == "Test Plan":
            testPlanID = issue_key
            tot_cnt_test_plan += 1
            testPlanType = ''
            if "fields.summary" in df_data.columns:
                summary = row['fields.summary'].split(" | ")
                if len(summary) > 2:
                    testPlanType = summary[1]  # get test plan type from summary
            test_case_cnts = get_test_cases_status(testPlanID)
            if testPlanType == 'Functional':
                cnt_functional_test_plan += 1
                cnt_auto_in_functional += test_case_cnts['cnt_auto']
                cnt_functional_test_cases += test_case_cnts['cnt_auto'] + test_case_cnts['cnt_manual']
            elif testPlanType == 'Regression':
                cnt_regression_test_plan += 1
                cnt_auto_in_regression += test_case_cnts['cnt_auto']
                cnt_regression_test_cases += test_case_cnts['cnt_auto'] + test_case_cnts['cnt_manual']

    # Prepare defect information in a single string with grouping by component
    defect_info = f"Total defects: {total_defects} Closed defects: {closed_defects} Open defects: {open_defects}"
    
    if open_defects > 0:
        defect_info += "\n\nOpen defects details:\n"
        for component, defects in open_defect_details_by_component.items():
            defect_info += f"\n{component}\n" + "\n".join(defects) + "\n"

    # Print or send defect information (if sending via email, ensure it's in HTML format)
    print(defect_info)





from bs4 import BeautifulSoup

def add_dark_mode_toggle(html_file_path):
    # Read the HTML file
    with open(html_file_path, 'r', encoding='utf-8') as file:
        soup = BeautifulSoup(file, 'html.parser')

    # Ensure the <head> tag exists
    if soup.head is None:
        soup.head = soup.new_tag("head")
        soup.insert(0, soup.head)

    # Ensure the <body> tag exists
    if soup.body is None:
        soup.body = soup.new_tag("body")
        soup.append(soup.body)
    
    # Add CSS for dark mode
    style_tag = soup.new_tag("style")
    style_tag.string = """
    body {
        background-color: white;
        color: black;
        transition: background-color 0.3s, color 0.3s;
    }

    /* Toggle Switch Styles */
    .toggle-switch {
        position: absolute;
        top: 20px;
        left: 50%;
        transform: translateX(-50%);
    }
    .switch {
        position: relative;
        display: inline-block;
        width: 40px;  /* Decreased width */
        height: 20px; /* Decreased height */
    }
    .switch input {
        opacity: 0;
        width: 0;
        height: 0;
    }
    .slider {
        position: absolute;
        cursor: pointer;
        top: 0;
        left: 0;
        right: 0;
        bottom: 0;
        background-color: #ccc;
        transition: 0.4s;
    }
    .slider:before {
        position: absolute;
        content: "";
        height: 16px;  /* Decreased height */
        width: 16px;   /* Decreased width */
        left: 2px;
        bottom: 2px;
        background-color: black;
        transition: 0.4s;
    }
    input:checked + .slider {
        background-color: green;
    }
    input:checked + .slider:before {
        transform: translateX(20px);  /* Adjusted based on the new switch width */
        background-color: white;
    }
    .slider.round {
        border-radius: 20px;  /* Adjusted based on new switch height */
    }
    .slider.round:before {
        border-radius: 50%;
    }

    /* Table Header Styles */
    th.white-purple {
        background-color: white;
        color: purple;
    }
    th.light-purple-navy {
        background-color: #E6E6FA; /* light purple */
        color: navy;
    }

    /* Table Cell Highlight Styles */
    td.yellow-highlight {
        background-color: yellow;
    }
    td.red-highlight {
        background-color: red;
        color: white;
    }

    /* Dark Mode Styles */
    body.dark-mode {
        background-color: #121212;
        color: white;
    }
    body.dark-mode th.white-purple {
        background-color: black;
        color: white;
    }
    body.dark-mode th.light-purple-navy {
        background-color: black;
        color: white;
    }
    body.dark-mode th.white-purple.header-5 {
        background-color: blue;
        color: white;
    }
    
    body.dark-mode td:not(.red-highlight):not(.yellow-highlight) {
        background-color: black;
        color: white;
    }
    
    """
    soup.head.append(style_tag)
    
    # Add toggle switch HTML
    toggle_switch_html = """
    <div class="toggle-switch">
        <label class="switch">
            <input type="checkbox" id="modeToggle">
            <span class="slider round"></span>
        </label>
    </div>
    """
    soup.body.insert(0, BeautifulSoup(toggle_switch_html, 'html.parser'))

    # Add JavaScript for toggling dark mode
    script_tag = soup.new_tag("script")
    script_tag.string = """
    const modeToggle = document.getElementById('modeToggle');
    modeToggle.addEventListener('change', () => {
        document.body.classList.toggle('dark-mode');
    });
    """
    soup.body.append(script_tag)

    # Save the modified HTML file
    with open("modified_" + html_file_path, 'w', encoding='utf-8') as file:
        file.write(str(soup))

    print(f"Dark mode toggle has been added to {html_file_path}. Output file is modified_{html_file_path}.")

# Example usage:
add_dark_mode_toggle("table.html")


'''
body.dark-mode td:not([style*="background-color: ##E6E6FA"]):not([style*="background-color: #ffff00"]) {
    background-color: black;
    color: white;
}
'''

body.dark-mode th[style*="background-color: #121212"] {
    background-color: #ffffff !important; /* New background color */
    color: #456736 !important; /* New text color */
}
